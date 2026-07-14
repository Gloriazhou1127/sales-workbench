"""
从花名册Excel导入用户到数据库
列映射: 1=姓名, 13=登录姓名, 14=角色, 15=L3负责人, 16=L4负责人, 17=初始密码
Org: col 2=一级组织, col 3=二级组织, col 4=三级组织, col 5=四级组织
col 6=职级, col 9=职务, col 1=工号(实际col 0 is the name?)
col 18=邮箱（可选）
"""
import sqlite3
import os
import sys
from werkzeug.security import generate_password_hash

# 加载环境变量
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))
except ImportError:
    pass

# Add server dir to path so we can use the DB_PATH
DB_PATH = os.path.join(os.path.dirname(__file__), 'workbench.db')
EXCEL_PATH = os.environ.get('ROSTER_EXCEL_PATH',
    r'C:\Work\M-企业事业部\2026系统问题\销售工作台\权限设置_花名册.xlsx')

def import_roster():
    try:
        import openpyxl
    except ImportError:
        print("需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    c = db.cursor()

    added = 0
    updated = 0
    errors = 0

    # Build leader maps first
    l3_leader_of = {}  # leader_name -> [subordinate_names]
    l4_leader_of = {}
    all_names = set()

    # First pass: collect all names and leader relationships
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        l3 = ws.cell(row=row_idx, column=15).value or ''
        l4 = ws.cell(row=row_idx, column=16).value or ''
        if not name:
            continue
        name = str(name).strip()
        all_names.add(name)
        l3 = str(l3).strip()
        l4 = str(l4).strip()
        if l3 and l3 != name:
            l3_leader_of.setdefault(l3, []).append(name)
        if l4 and l4 != name:
            l4_leader_of.setdefault(l4, []).append(name)

    # Second pass: import users
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if not name:
            continue
        name = str(name).strip()
        login_name = ws.cell(row=row_idx, column=13).value or ''
        login_name = str(login_name).strip()
        role_str = str(ws.cell(row=row_idx, column=14).value or '')
        l3 = str(ws.cell(row=row_idx, column=15).value or '').strip()
        l4 = str(ws.cell(row=row_idx, column=16).value or '').strip()
        pwd_raw = str(ws.cell(row=row_idx, column=17).value or '')
        # Use the first column's value - col 2 might be job number
        job_num_raw = ws.cell(row=row_idx, column=2).value or ''
        job_num = str(job_num_raw).strip()
        dept_l1 = str(ws.cell(row=row_idx, column=3).value or '').strip() if ws.cell(row=row_idx, column=3).value else ''
        dept_l2 = str(ws.cell(row=row_idx, column=4).value or '').strip() if ws.cell(row=row_idx, column=4).value else ''
        dept_l3 = str(ws.cell(row=row_idx, column=5).value or '').strip() if ws.cell(row=row_idx, column=5).value else ''
        dept_l4 = str(ws.cell(row=row_idx, column=6).value or '').strip() if ws.cell(row=row_idx, column=6).value else ''
        grade = str(ws.cell(row=row_idx, column=7).value or '').strip() if ws.cell(row=row_idx, column=7).value else ''
        job_title = str(ws.cell(row=row_idx, column=10).value or '').strip() if ws.cell(row=row_idx, column=10).value else ''
        email_raw = str(ws.cell(row=row_idx, column=18).value or '') if ws.cell(row=row_idx, column=18).value else ''
        email = email_raw.strip() if email_raw != 'None' else ''

        if not login_name:
            print(f"  ⚠ Row {row_idx}: 跳过 {name}, 无登录姓名")
            errors += 1
            continue

        # Determine roles from the 角色 column
        roles_list = [r.strip() for r in role_str.split(',') if r.strip()]
        is_admin = '管理员' in roles_list
        is_city = '城市管理' in roles_list
        is_sales = '销售顾问' in roles_list

        # Default role for backward compatibility
        if is_admin:
            base_role = 'admin'
        elif is_city:
            base_role = 'user'  # 城市管理 = L4负责人权限
        elif is_sales:
            base_role = 'user'
        else:
            base_role = 'user'

        # Determine leader flags
        is_l3_leader = 1 if name in l3_leader_of and len(l3_leader_of[name]) > 0 else 0
        is_l4_leader = 1 if name in l4_leader_of and len(l4_leader_of[name]) > 0 else 0

        # For multi-role storage
        multi_roles = ','.join(roles_list)

        if not pwd_raw or pwd_raw == 'None':
            pwd_hash = generate_password_hash('123456')
        else:
            pwd_hash = generate_password_hash(pwd_raw)

        # Check if user exists
        existing = c.execute("SELECT id FROM users WHERE username = ?", (login_name,)).fetchone()
        if existing:
            c.execute("""
                UPDATE users SET
                    display_name=?, role=?, job_number=?, dept_l3=?, dept_l4=?,
                    l3_leader=?, l4_leader=?, is_l3_leader=?, is_l4_leader=?,
                    grade=?, job_title=?, roles=?, email=?
                WHERE username=?
            """, (name, base_role, job_num, dept_l3, dept_l4,
                  l3, l4, is_l3_leader, is_l4_leader,
                  grade, job_title, multi_roles, email, login_name))
            updated += 1
            print(f"  ✓ Updated: {name} ({login_name}) role={multi_roles}")
        else:
            c.execute("""
                INSERT INTO users (username, password, display_name, role, job_number,
                    dept_l3, dept_l4, l3_leader, l4_leader, is_l3_leader, is_l4_leader,
                    grade, job_title, roles, enabled, email)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            """, (login_name, pwd_hash, name, base_role, job_num,
                  dept_l3, dept_l4, l3, l4, is_l3_leader, is_l4_leader,
                  grade, job_title, multi_roles, email))
            added += 1
            print(f"  + Added: {name} ({login_name}) role={multi_roles} L3ldr={is_l3_leader} L4ldr={is_l4_leader}")

    # Clear and rebuild org_hierarchy
    c.execute("DELETE FROM org_hierarchy")
    for leader, subs in l3_leader_of.items():
        for sub in subs:
            c.execute("""
                INSERT OR IGNORE INTO org_hierarchy (leader_name, level, subordinate_name, dept_l3, dept_l4)
                VALUES (?, 'L3', ?, '', '')
            """, (leader, sub))
    for leader, subs in l4_leader_of.items():
        for sub in subs:
            c.execute("""
                INSERT OR IGNORE INTO org_hierarchy (leader_name, level, subordinate_name, dept_l3, dept_l4)
                VALUES (?, 'L4', ?, '', '')
            """, (leader, sub))

    db.commit()

    # Print summary
    print(f"\n{'='*50}")
    print(f"导入完成: 新增 {added} 人, 更新 {updated} 人, 错误 {errors} 人")
    print(f"三级组织负责人: {len(l3_leader_of)} 人")
    print(f"四级组织负责人: {len(l4_leader_of)} 人")
    
    total = c.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    l3_count = c.execute("SELECT COUNT(*) FROM users WHERE is_l3_leader=1").fetchone()[0]
    l4_count = c.execute("SELECT COUNT(*) FROM users WHERE is_l4_leader=1").fetchone()[0]
    org_count = c.execute("SELECT COUNT(*) FROM org_hierarchy").fetchone()[0]
    print(f"总用户数: {total} | L3负责人: {l3_count} | L4负责人: {l4_count} | 组织关系: {org_count}")
    
    db.close()
    return added, updated, errors

if __name__ == '__main__':
    import_roster()
