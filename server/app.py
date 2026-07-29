"""
销售工作台 - Flask 后端
提供 JWT 鉴权 + 模块链接后台管理
"""
import os
import json
import sqlite3
import datetime as dt
import secrets
from functools import wraps

from flask import Flask, request, jsonify, g, send_from_directory
from flask_cors import CORS
import jwt
from werkzeug.security import generate_password_hash, check_password_hash

# 加载 .env 文件（生产环境优先从系统环境变量读取）
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))
except ImportError:
    pass

# ============================================================
# 配置（全部从环境变量读取）
# ============================================================
SECRET_KEY = os.environ.get('SECRET_KEY', secrets.token_hex(32))
DB_PATH = os.path.join(os.path.dirname(__file__),
                       os.environ.get('DATABASE_PATH', 'workbench.db'))
IS_PRODUCTION = os.environ.get('FLASK_ENV', 'development') == 'production'

app = Flask(__name__, static_folder='../static', static_url_path='')

# CORS 限制：生产环境限定域名，开发环境放通
if IS_PRODUCTION:
    cors_origins = os.environ.get('CORS_ORIGINS', '')
    if cors_origins:
        origins = [o.strip() for o in cors_origins.split(',') if o.strip()]
        CORS(app, origins=origins, supports_credentials=True)
    else:
        CORS(app)
else:
    CORS(app)

# ============================================================
# 数据库初始化
# ============================================================

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db:
        db.close()

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    c = db.cursor()

    # 用户表（含权限和组织信息）
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        display_name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'user',
        job_number TEXT,
        dept_l3 TEXT,
        dept_l4 TEXT,
        l3_leader TEXT,
        l4_leader TEXT,
        is_l3_leader INTEGER DEFAULT 0,
        is_l4_leader INTEGER DEFAULT 0,
        grade TEXT,
        job_title TEXT,
        roles TEXT DEFAULT '',
        email TEXT,
        enabled INTEGER DEFAULT 1
    )''')

    # 组织层级关系表
    c.execute('''CREATE TABLE IF NOT EXISTS org_hierarchy (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        leader_name TEXT NOT NULL,
        level TEXT NOT NULL,
        subordinate_name TEXT NOT NULL,
        dept_l3 TEXT,
        dept_l4 TEXT,
        UNIQUE(leader_name, level, subordinate_name)
    )''')

    # 跨三级组织可见性：某负责人可额外查看指定三级部门的数据（花名册重导不清除）
    c.execute('''CREATE TABLE IF NOT EXISTS cross_dept_visibility (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        leader_name TEXT NOT NULL,
        view_dept_l3 TEXT NOT NULL,
        UNIQUE(leader_name, view_dept_l3)
    )''')

    # 数据导入日志表
    c.execute('''CREATE TABLE IF NOT EXISTS data_import_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_name TEXT,
        data_type TEXT,
        row_count INTEGER,
        success_count INTEGER,
        error_count INTEGER,
        filtered_count INTEGER DEFAULT 0,
        status TEXT DEFAULT 'done',
        imported_by TEXT,
        imported_at TEXT DEFAULT (datetime('now','localtime'))
    )''')
    # 兼容旧表：添加缺失列
    try:
        c.execute('ALTER TABLE data_import_log ADD COLUMN filtered_count INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass

    # 模块配置表
    c.execute('''CREATE TABLE IF NOT EXISTS modules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT,
        link_url TEXT,
        link_type TEXT DEFAULT 'online',
        icon_svg TEXT,
        icon_color TEXT DEFAULT '#185FA5',
        icon_bg TEXT DEFAULT '#E6F1FB',
        visible_to TEXT DEFAULT 'all',
        sort_order INTEGER DEFAULT 0,
        enabled INTEGER DEFAULT 1,
        updated_at TEXT
    )''')

    # 插入默认管理员
    c.execute("SELECT id FROM users WHERE username = 'admin'")
    if not c.fetchone():
        admin_pwd = os.environ.get('ADMIN_DEFAULT_PASSWORD', 'admin123')
        c.execute("INSERT INTO users (username, password, display_name, role) VALUES (?, ?, ?, ?)",
                  ('admin', generate_password_hash(admin_pwd), '管理员', 'admin'))

    # 插入默认普通用户
    c.execute("SELECT id FROM users WHERE username = 'sales'")
    if not c.fetchone():
        sales_pwd = os.environ.get('SALES_DEFAULT_PASSWORD', 'sales123')
        c.execute("INSERT INTO users (username, password, display_name, role) VALUES (?, ?, ?, ?)",
                  ('sales', generate_password_hash(sales_pwd), '销售小王', 'user'))

    # 插入默认模块
    c.execute("SELECT COUNT(*) FROM modules")
    if c.fetchone()[0] == 0:
        default_modules = [
            ('OA', '审批、流程、日常办公入口', 'https://oa.example.com', 'online',
             '<rect x=3 y=3 width=18 height=18 rx=2/><path d=M3 9h18M9 21V9/>', '#185FA5', '#E6F1FB', 0),
            ('报价单', '创建与管理客户报价方案', 'http://localhost:5173', 'local',
             '<path d=M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z/><polyline points=14 2 14 8 20 8/>', '#854F0B', '#FAEEDA', 1),
            ('销售易', 'CRM 客户管理与商机跟踪', 'https://crm.example.com', 'online',
             '<path d=M20 7H4a2 2 0 00-2 2v6a2 2 0 002 2h16a2 2 0 002-2V9a2 2 0 00-2-2z/><circle cx=12 cy=12 r=2/>', '#534AB7', '#EEEDFE', 2),
            ('目标管理', '销售目标设置与进度追踪', 'http://localhost:5000/dashboard', 'local',
             '<polyline points=22 12 18 12 15 21 9 3 6 12 2 12/>', '#3B6D11', '#EAF3DE', 3),
            ('新产品推进', '新品落地进度与任务跟踪', 'http://localhost:8080', 'local',
             '<polygon points=12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2/>', '#993C1D', '#FAECE7', 4),
            ('重点客户跟进', '重点客户拜访与商机记录', '', 'todo',
             '<path d=M17 21v-2a4 4 0 00-4-4H5a4 4 0 00-4 4v2/><circle cx=9 cy=7 r=4/><path d=M23 21v-2a4 4 0 00-3-3.87/><path d=M16 3.13a4 4 0 010 7.75/>', '#0F6E56', '#E1F5EE', 5),
            ('市场活动跟进', '活动策划、执行与效果复盘', '', 'todo',
             '<rect x=3 y=4 width=18 height=18 rx=2 ry=2/><line x1=16 y1=2 x2=16 y2=6/><line x1=8 y1=2 x2=8 y2=6/><line x1=3 y1=10 x2=21 y2=10/>', '#854F0B', '#FAEEDA', 6),
            ('权限管理', '账号与角色权限配置', '', 'admin',
             '<rect x=3 y=11 width=18 height=11 rx=2 ry=2/><path d=M7 11V7a5 5 0 0110 0v4/>', '#993556', '#FBEAF0', 7,
             'admin'),
            ('提成系统', '销售提成计算与发放管理', '', 'todo',
             '<path d=M12 2L15.09 8.26L22 9.27L17 14.14L18.18 21.02L12 17.77L5.82 21.02L7 14.14L2 9.27L8.91 8.26L12 2Z/>', '#854F0B', '#FAEEDA', 8),
            ('硬件售卖地图', '区域硬件销售分布与热力图', '', 'todo',
             '<circle cx=12 cy=12 r=10/><polygon points=16.24 7.76 14.12 14.12 7.76 16.24 9.88 9.88 16.24 7.76/>', '#0F6E56', '#E1F5EE', 9),
            ('销售日报', '今日销售活动汇总与周进度追踪', '/daily-report', 'local',
             '<rect x=3 y=4 width=18 height=18 rx=2/><path d=M16 2v4M8 2v4M3 10h18M8 14h.01M12 14h.01M16 14h.01M8 18h.01M12 18h.01M16 18h.01/>', '#3B6D11', '#EAF3DE', 10),
        ]
        for idx, m in enumerate(default_modules):
            if len(m) == 9:
                name, desc, url, ltype, icon, color, bg, sort, visible = m
            else:
                name, desc, url, ltype, icon, color, bg, sort = m
                visible = 'all' if ltype != 'admin' else 'admin'
            c.execute('''INSERT INTO modules (name, description, link_url, link_type, icon_svg, icon_color, icon_bg, visible_to, sort_order, updated_at)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (name, desc, url, ltype, icon, color, bg, visible, sort, dt.datetime.now().isoformat()))

    db.commit()

    # 兼容已有数据库：自动新增缺失列
    try:
        c = db.cursor()
        cols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
        if 'email' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN email TEXT")
            print("[init_db] 已新增 users.email 列")
        # key_account_hardware 新增暂停跟进字段
        hw_cols = [r[1] for r in c.execute("PRAGMA table_info(key_account_hardware)").fetchall()]
        for col_name in ['is_paused', 'pause_reason', 'paused_by', 'paused_at']:
            if col_name not in hw_cols:
                col_type = 'INTEGER DEFAULT 0' if col_name == 'is_paused' else 'TEXT'
                db.execute(f"ALTER TABLE key_account_hardware ADD COLUMN {col_name} {col_type}")
        # 兼容修复：确保所有NULL的is_paused变为0（每次启动都执行，幂等安全）
        db.execute("UPDATE key_account_hardware SET is_paused = 0 WHERE is_paused IS NULL")
        print("[init_db] key_account_hardware 暂停字段检查完成")
    except Exception as e:
        print(f"[init_db] 列迁移跳过: {e}")

    db.close()

# ============================================================
# JWT 鉴权
# ============================================================

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return jsonify({'error': '未登录'}), 401
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
            g.current_user = payload
        except jwt.ExpiredSignatureError:
            return jsonify({'error': '登录已过期'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': '无效的登录凭证'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if g.current_user.get('role') != 'admin':
            return jsonify({'error': '权限不足，仅管理员可操作'}), 403
        return f(*args, **kwargs)
    return decorated

# ============================================================
# 权限过滤辅助函数
# ============================================================

def get_visible_owners(db, current_user):
    """
    根据当前用户角色返回可见的 owner_name 列表。
    - admin: 返回 None (看全部)
    - L3 负责人: 自己 + 所有L3下属
    - L4 负责人/城市管理: 自己 + 所有L4下属
    - 普通销售顾问: 仅自己
    """
    role = current_user.get('role', 'user')
    display_name = current_user.get('display_name', '')

    # admin 看全部
    if role == 'admin':
        return None

    is_l3 = current_user.get('is_l3_leader', 0)
    is_l4 = current_user.get('is_l4_leader', 0)

    # 从数据库查询完整的用户信息（含roles字段）
    user = db.execute(
        "SELECT roles, is_l3_leader, is_l4_leader FROM users WHERE username = ?",
        (current_user.get('username', ''),)
    ).fetchone()
    if not user:
        return [display_name]

    roles_str = user['roles'] or ''
    is_l3_db = user['is_l3_leader'] or 0
    is_l4_db = user['is_l4_leader'] or 0

    owners = {display_name}

    # L3 负责人：查组织层级
    if is_l3_db:
        l3_subs = db.execute(
            "SELECT subordinate_name FROM org_hierarchy WHERE leader_name = ? AND level = 'L3'",
            (display_name,)
        ).fetchall()
        for s in l3_subs:
            owners.add(s['subordinate_name'])

    # L4 负责人：查组织层级
    if is_l4_db:
        l4_subs = db.execute(
            "SELECT subordinate_name FROM org_hierarchy WHERE leader_name = ? AND level = 'L4'",
            (display_name,)
        ).fetchall()
        for s in l4_subs:
            owners.add(s['subordinate_name'])

    # 也检查 roles 中的 城市管理 等同于 L4
    if '城市管理' in roles_str and not is_l4_db:
        l4_subs = db.execute(
            "SELECT subordinate_name FROM org_hierarchy WHERE leader_name = ? AND level = 'L4'",
            (display_name,)
        ).fetchall()
        for s in l4_subs:
            owners.add(s['subordinate_name'])

    # 跨三级组织可见性：额外查看指定三级部门的数据（持久，不受花名册重导影响）
    cross_rows = db.execute(
        "SELECT view_dept_l3 FROM cross_dept_visibility WHERE leader_name = ?",
        (display_name,)
    ).fetchall()
    for cr in cross_rows:
        dept = cr['view_dept_l3']
        dep_rows = db.execute(
            "SELECT DISTINCT owner_name FROM key_account_hardware "
            "WHERE department = ? AND owner_name IS NOT NULL AND owner_name != ''",
            (dept,)
        ).fetchall()
        for o in dep_rows:
            own = o['owner_name']
            prefix = own.split('-', 1)[0] if '-' in own else own
            if prefix:
                owners.add(prefix)

    return list(owners)


def get_owner_filter(db, current_user, table_alias='owner_name'):
    """
    返回 (where_clause, params) 用于 SQL 权限过滤。
    管理员返回 ("", [])，其他返回 ("(owner_name LIKE ? OR owner_name LIKE ? ...)", [...])。
    注意：owner_name 格式为 "姓名-工号"，display_name 为 "姓名"，用 LIKE 前缀匹配。
    """
    owners = get_visible_owners(db, current_user)
    if owners is None:
        return '', []
    # 使用 LIKE 前缀匹配，因为 owner_name 是 "姓名-工号" 格式
    like_clauses = ' OR '.join([f"{table_alias} LIKE ?" for _ in owners])
    like_patterns = [f"{name}-%" for name in owners]
    return f"({like_clauses})", like_patterns


# ============================================================
# API 路由
# ============================================================

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()

    if not user or not check_password_hash(user['password'], password):
        return jsonify({'error': '用户名或密码错误'}), 401

    # 离职/停用账号拦截
    if user.get('enabled', 1) != 1:
        return jsonify({'error': '账号已停用，请联系管理员'}), 403

    payload = {
        'id': user['id'],
        'username': user['username'],
        'display_name': user['display_name'],
        'role': user['role'],
        'is_l3_leader': user['is_l3_leader'] if 'is_l3_leader' in user.keys() else 0,
        'is_l4_leader': user['is_l4_leader'] if 'is_l4_leader' in user.keys() else 0,
        'exp': dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=24)
    }
    token = jwt.encode(payload, SECRET_KEY, algorithm='HS256')

    return jsonify({
        'token': token,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'display_name': user['display_name'],
            'role': user['role']
        }
    })

@app.route('/api/users/by-role', methods=['GET'])
def list_users_by_role():
    """按角色返回可用用户名列表（登录页使用）"""
    role = request.args.get('role', '')
    db = get_db()

    if role == '管理员':
        rows = db.execute(
            "SELECT username, display_name FROM users WHERE role = 'admin' OR roles LIKE ? AND enabled = 1 ORDER BY username",
            ('%管理员%',)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT username, display_name FROM users WHERE roles LIKE ? AND enabled = 1 ORDER BY username",
            (f'%{role}%',)
        ).fetchall()

    return jsonify({'users': [{'username': r['username'], 'display_name': r['display_name']} for r in rows]})


@app.route('/api/user/me', methods=['GET'])
@token_required
def get_me():
    return jsonify(g.current_user)


@app.route('/api/user/change-password', methods=['POST'])
@token_required
def change_password():
    """当前用户修改自己的密码"""
    data = request.get_json()
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')

    if not old_password or not new_password:
        return jsonify({'error': '旧密码和新密码不能为空'}), 400
    if len(new_password) < 4:
        return jsonify({'error': '新密码至少4位'}), 400
    if old_password == new_password:
        return jsonify({'error': '新密码不能与旧密码相同'}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (g.current_user['id'],)).fetchone()
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    if not check_password_hash(user['password'], old_password):
        return jsonify({'error': '旧密码错误'}), 401

    db.execute("UPDATE users SET password = ? WHERE id = ?",
               (generate_password_hash(new_password), g.current_user['id']))
    db.commit()
    return jsonify({'success': True, 'message': '密码修改成功'})


# ============================================================
# 邮件发送工具
# ============================================================
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.exmail.qq.com')   # 企业微信邮箱默认
SMTP_PORT = int(os.environ.get('SMTP_PORT', 465))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASS = os.environ.get('SMTP_PASS', '')
SMTP_FROM_NAME = os.environ.get('SMTP_FROM_NAME', '销售工作台')
SMTP_USE_SSL = os.environ.get('SMTP_SSL', 'true').lower() == 'true'

def send_email(to_addr, subject, html_body):
    """通过 SMTP 发送邮件，成功返回 None，失败返回错误信息"""
    if not SMTP_USER or not SMTP_PASS:
        return '邮件服务未配置（缺少 SMTP_USER / SMTP_PASS）'
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = f'{SMTP_FROM_NAME} <{SMTP_USER}>'
        msg['To'] = to_addr
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        if SMTP_USE_SSL:
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15)
        else:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15)
            server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, [to_addr], msg.as_string())
        server.quit()
        return None
    except Exception as e:
        return str(e)


@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    """忘记密码：通过公司邮箱重置密码"""
    data = request.get_json()
    email = (data.get('email') or '').strip().lower()

    if not email:
        return jsonify({'error': '请输入邮箱地址'}), 400
    if '@' not in email:
        return jsonify({'error': '邮箱格式不正确'}), 400

    db = get_db()
    user = db.execute("SELECT id, username, display_name FROM users WHERE LOWER(email) = ? AND enabled = 1", (email,)).fetchone()
    db.close()

    if not user:
        # 安全考虑：不暴露是否存在该用户
        return jsonify({
            'success': True,
            'message': '如果该邮箱已注册，新密码将发送至您的邮箱（请在垃圾邮件中检查）'
        })

    # 生成随机8位密码
    new_pwd = secrets.token_urlsafe(6)

    # 更新数据库中的密码
    db = get_db()
    db.execute("UPDATE users SET password = ? WHERE id = ?",
               (generate_password_hash(new_pwd), user['id']))
    db.commit()
    db.close()

    # 发送邮件
    html_body = f"""<div style="font-family: -apple-system,'PingFang SC','Microsoft YaHei',sans-serif; max-width:560px; margin:0 auto; padding:30px 20px;">
<h2 style="color:#1a1a1a;">销售工作台 — 密码重置</h2>
<p style="color:#555;font-size:14px;line-height:1.8;">
您好，<strong>{user['display_name']}</strong>！<br>
您申请了密码重置操作。以下是您的新登录密码：
</p>
<div style="background:#f0f7ff;border:1px solid #d0e3ff;border-radius:10px;padding:20px;text-align:center;margin:18px 0;">
<p style="font-size:13px;color:#666;margin-bottom:8px;">新密码</p>
<p style="font-size:24px;font-weight:700;color:#185FA5;letter-spacing:2px;">{new_pwd}</p>
</div>
<p style="color:#888;font-size:12px;line-height:1.8;">
登录后请及时在「工作台 → 修改密码」中更换为便于记忆的密码。<br>
此邮件由系统自动发送，请勿回复。
</p>
<p style="margin-top:24px;font-size:12px;color:#bbb;">
— 销售工作台 · {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}
</p>
</div>"""

    send_err = send_email(email, '【销售工作台】密码重置通知', html_body)

    if send_err:
        print(f"[forgot-password] 邮件发送失败({user['username']}→{email}): {send_err}")
        return jsonify({
            'success': True,
            'message': '处理中...如果未收到邮件，请联系管理员'
        })

    print(f"[forgot-password] 密码已重置并发送至 {user['username']}→{email}")
    return jsonify({
        'success': True,
        'message': '新密码已发送至您的邮箱，请注意查收（含垃圾邮件箱）'
    })


@app.route('/api/modules', methods=['GET'])
@token_required
def list_modules():
    role = g.current_user.get('role', 'user')
    db = get_db()
    rows = db.execute(
        "SELECT * FROM modules WHERE enabled = 1 ORDER BY sort_order"
    ).fetchall()

    modules = []
    for r in rows:
        # admin 类型的模块仅管理员可见
        if r['visible_to'] == 'admin' and role != 'admin':
            continue
        modules.append({
            'id': r['id'],
            'name': r['name'],
            'description': r['description'],
            'link_url': r['link_url'],
            'link_type': r['link_type'],
            'icon_svg': r['icon_svg'],
            'icon_color': r['icon_color'],
            'icon_bg': r['icon_bg'],
            'visible_to': r['visible_to'],
            'sort_order': r['sort_order'],
        })

    return jsonify({'modules': modules})

@app.route('/api/modules/<int:module_id>', methods=['PUT'])
@token_required
@admin_required
def update_module(module_id):
    data = request.get_json()
    db = get_db()

    allowed_fields = ['name', 'description', 'link_url', 'link_type',
                      'icon_color', 'icon_bg', 'visible_to', 'sort_order', 'enabled']
    updates = {}
    for k in allowed_fields:
        if k in data:
            updates[k] = data[k]

    if not updates:
        return jsonify({'error': '没有可更新的字段'}), 400

    set_clause = ', '.join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [dt.datetime.now().isoformat(), module_id]
    set_clause += ", updated_at = ?"

    db.execute(f"UPDATE modules SET {set_clause} WHERE id = ?", values)
    db.commit()

    return jsonify({'message': '更新成功'})

@app.route('/api/modules/<int:module_id>', methods=['GET'])
@token_required
@admin_required
def get_module_detail(module_id):
    db = get_db()
    r = db.execute("SELECT * FROM modules WHERE id = ?", (module_id,)).fetchone()
    if not r:
        return jsonify({'error': '模块不存在'}), 404
    return jsonify({
        'id': r['id'],
        'name': r['name'],
        'description': r['description'],
        'link_url': r['link_url'],
        'link_type': r['link_type'],
        'icon_svg': r['icon_svg'],
        'icon_color': r['icon_color'],
        'icon_bg': r['icon_bg'],
        'visible_to': r['visible_to'],
        'sort_order': r['sort_order'],
        'enabled': r['enabled'],
    })

# ============================================================
# 重点客户跟进 — 数据库初始化（追加到 init_db 之后调用）
# ============================================================

def init_key_account_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    c = db.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS key_account_hardware (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        crm_account_id TEXT UNIQUE,          -- 销售易客户 ID
        customer_name  TEXT,                 -- 客户名称
        department     TEXT,                 -- 所属部门
        owner_name     TEXT,                 -- 客户所有人
    created_date   TEXT,                 -- 客户创建日期（来自CRM清单）
    lent_date      TEXT,                 -- 借用日期（最早订单日期）
    purchase_date  TEXT,                 -- 采购日期
    checkin_date   TEXT,                 -- 打点日期 = MIN(lent_date, purchase_date)
    classroom_count INTEGER DEFAULT 0,  -- 教室数量
    lent_units     INTEGER DEFAULT 0,   -- 累计借用台数
    purchased_units INTEGER DEFAULT 0,  -- 累计采购台数
        visit_count    INTEGER DEFAULT 0,   -- 累计拜访次数
        last_visit_date TEXT,               -- 最近拜访日期（距今天数由前端计算）
        opportunity_count  INTEGER DEFAULT 0,-- 商机数量
        opportunity_amount REAL DEFAULT 0,   -- 商机金额（元）
        opportunity_status TEXT,             -- 商机状态（如：方案中/报价中/谈判中/已成交/已丢单）

        -- 顾问手动填写字段
        is_checked_in  INTEGER DEFAULT 0,   -- 是否已打点 0/1
        checkin_stage  TEXT,                -- 打点阶段 1-6
        stage_desc     TEXT,                -- 阶段描述
        feedback       TEXT,                -- 客户现阶段反馈
        has_blocker    INTEGER DEFAULT 0,   -- 是否有卡点 0/1
        blocker_detail TEXT,                -- 卡点描述

        -- 阶段4：整体复盘 & 价值量化
        stage4_reported       INTEGER DEFAULT 0, -- 是否已汇报给决策人
        stage4_decision_role  TEXT,               -- 决策人身份
        stage4_meeting_held   INTEGER DEFAULT 0, -- 是否已组织家长会
        stage4_meeting_files  TEXT,               -- 家长会附件(JSON数组)

        -- 阶段6：无采购意向
        stage6_no_intent_reason TEXT,           -- 无采购意向原因

        -- 暂停跟进
        is_paused     INTEGER DEFAULT 0,       -- 是否暂停跟进 0/1
        pause_reason  TEXT,                     -- 暂停原因
        paused_by     TEXT,                     -- 操作人
        paused_at     TEXT,                     -- 暂停时间

        updated_by     TEXT,                -- 最后更新人
        updated_at     TEXT,                -- 最后更新时间
        synced_at      TEXT                 -- 最后同步时间
    )''')
    db.commit()
    db.close()

def init_stages_config():
    """管理员可配置的打点阶段定义"""
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    c = db.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS stages_config (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stage_key TEXT UNIQUE NOT NULL,
        stage_name TEXT NOT NULL,
        stage_desc TEXT,
        sort_order INTEGER DEFAULT 0
    )''')
    defaults = [
        ('1', '① 零星采购 OR 推进试用', '客户有零星采购或正在推进试用', 1),
        ('2', '② 落地使用 & 首轮跟进', '产品已落地使用，进行首轮跟进', 2),
        ('3', '③ 深度应用 & 每周运维', '深度使用中，每周运维跟进', 3),
        ('4', '④ 整体复盘 & 价值量化', '整体复盘，量化客户价值', 4),
        ('5', '⑤ 政策推送 & 促成集中采购', '政策推送中，推动集中采购', 5),
        ('6', '⑥ 客户无采购意向', '客户暂无采购意向', 6),
    ]
    for sk, sn, sd, so in defaults:
        c.execute(
            "INSERT OR IGNORE INTO stages_config (stage_key, stage_name, stage_desc, sort_order) VALUES (?, ?, ?, ?)",
            (sk, sn, sd, so)
        )
    db.commit()
    db.close()

def init_followup_logs():
    """每周填报历史记录表"""
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute('''CREATE TABLE IF NOT EXISTS key_account_followup_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hardware_id   INTEGER NOT NULL,
        fill_round    TEXT NOT NULL,
        fill_date     TEXT,
        checkin_stage TEXT,
        stage_desc    TEXT,
        feedback      TEXT,
        has_blocker   INTEGER DEFAULT 0,
        blocker_detail TEXT,
        followup_plan  TEXT,
        opportunity_status TEXT,
        stage4_reported      INTEGER DEFAULT 0,
        stage4_decision_role TEXT,
        stage4_meeting_held  INTEGER DEFAULT 0,
        stage4_meeting_files TEXT,
        stage6_no_intent_reason TEXT,
        filled_by     TEXT,
        filled_at     TEXT,
        updated_at    TEXT,
        UNIQUE(hardware_id, fill_round),
        FOREIGN KEY (hardware_id) REFERENCES key_account_hardware(id)
    )''')
    db.execute("CREATE INDEX IF NOT EXISTS idx_followup_hw ON key_account_followup_logs(hardware_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_followup_round ON key_account_followup_logs(fill_round)")

    # 重点客户跟进页面访问记录
    db.execute('''CREATE TABLE IF NOT EXISTS key_account_visits (
        user_id INTEGER PRIMARY KEY,
        last_visited_at TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    db.commit()
    db.close()

# ============================================================
@app.route('/api/key-account/reminder', methods=['GET'])
@token_required
def ka_reminder():
    """待办提醒：未填报数 + 截止日期（仅当前用户可见范围）"""
    db = get_db()
    current_round = _current_fill_round()

    # 权限过滤：只看自己能看到的客户
    owner_where, owner_params = get_owner_filter(db, g.current_user)

    # 截止日期 = 本周日（轮次本身就是周日）
    current_sunday = dt.date.fromisoformat(current_round)
    deadline_date = current_sunday  # Sunday
    today = dt.date.today()
    days_remaining = max(0, (deadline_date - today).days)

    # 本周已填报的记录数（周一~周日区间，带权限过滤）
    wk_m, wk_s = _week_range(current_round)
    if owner_where:
        filled_sql = f"""SELECT COUNT(DISTINCT f.hardware_id) FROM key_account_followup_logs f
                         INNER JOIN key_account_hardware h ON h.id = f.hardware_id
                         WHERE f.fill_round>=? AND f.fill_round<=? AND COALESCE(h.is_paused,0)=0 AND h.{owner_where}"""
        filled_count = db.execute(filled_sql, [wk_m, wk_s] + owner_params).fetchone()[0]

        total_count = db.execute(
            f"SELECT COUNT(*) FROM key_account_hardware WHERE COALESCE(is_paused,0)=0 AND {owner_where}", owner_params
        ).fetchone()[0]
    else:
        # 管理员看全部（排除暂停）
        filled_count = db.execute(
            "SELECT COUNT(DISTINCT hardware_id) FROM key_account_followup_logs WHERE fill_round>=? AND fill_round<=?",
            (wk_m, wk_s)
        ).fetchone()[0]
        total_count = db.execute("SELECT COUNT(*) FROM key_account_hardware WHERE COALESCE(is_paused,0)=0").fetchone()[0]

    unfilled_count = max(0, total_count - filled_count)

    return jsonify({
        'fill_round': current_round,
        'deadline_date': deadline_date.isoformat(),
        'days_remaining': days_remaining,
        'total_count': total_count,
        'filled_count': filled_count,
        'unfilled_count': unfilled_count,
    })

@app.route('/api/key-account/visit', methods=['POST'])
@token_required
def ka_record_visit():
    """记录用户访问重点客户跟进页面"""
    db = get_db()
    user_id = g.current_user['id']
    db.execute('''INSERT OR REPLACE INTO key_account_visits (user_id, last_visited_at)
                  VALUES (?, datetime('now','localtime'))''', (user_id,))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/admin/key-account-visits', methods=['GET'])
@token_required
def admin_key_account_visits():
    """管理员查看所有用户对重点客户跟进的访问记录"""
    if g.current_user.get('role') != 'admin':
        return jsonify({'error': '无权限'}), 403
    db = get_db()
    rows = db.execute('''
        SELECT u.id, u.username, u.display_name, u.role,
               COALESCE(u.dept_l4, u.dept_l3) AS department,
               v.last_visited_at
        FROM users u
        LEFT JOIN key_account_visits v ON v.user_id = u.id
        WHERE u.role != 'admin'
        ORDER BY v.last_visited_at DESC NULLS LAST, u.display_name ASC
    ''').fetchall()
    result = []
    for r in rows:
        result.append({
            'id': r['id'],
            'username': r['username'],
            'display_name': r['display_name'],
            'role': r['role'],
            'department': r['department'],
            'last_visited_at': r['last_visited_at'],
        })
    return jsonify({'users': result})

# 重点客户跟进 — API
# ============================================================

import urllib.request
import urllib.error

CRM_BI_VIEWS = {
    'customers':    '4370506283467544',   # 客户清单（教室数）
    'lent':         '4370186273653472',   # 借用明细
    'sold':         '4370146450899739',   # 售卖明细（采购数量）
    'visits':       '4368585711944422',   # 拜访明细
    'opportunity':  '4182163232883411',   # 商机复盘表
}


def _crm_view_query(view_id, page=1, page_size=500):
    """通过 MCP 内部 HTTP 桥调用 bi-viewQuery（如可用）；
    此处以占位形式返回 None，实际同步走 /api/key-account/sync 时由前端触发并由 MCP 客户端代理。
    """
    return None


@app.route('/api/key-account/hardware', methods=['GET'])
@token_required
def ka_hardware_list():
    """返回硬件打点跟进列表"""
    db = get_db()

    # 筛选参数
    stage   = request.args.get('stage', '')
    checked = request.args.get('checked', '')   # '1'=已打点 '0'=未打点
    filled  = request.args.get('filled', '')    # '1'=本周已填报 '0'=本周未填报
    dept    = request.args.get('dept', '')
    search  = request.args.get('q', '')
    paused  = request.args.get('paused', '')     # ''=仅跟进中 '1'=仅暂停 'all'=全部
    page    = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))

    where_clauses = []
    params = []

    # 暂停状态过滤：默认只看跟进中（兼容NULL值）
    if paused == '1':
        where_clauses.append("is_paused = 1")
    elif paused != 'all':
        where_clauses.append("COALESCE(is_paused, 0) = 0")

    # 权限过滤：根据角色+组织层级返回可见的 owner_name 列表
    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')
    if owner_filter:
        where_clauses.append(owner_filter)
        params.extend(owner_params)

    if stage:
        where_clauses.append("checkin_stage = ?")
        params.append(stage)
    if checked in ('0', '1'):
        where_clauses.append("is_checked_in = ?")
        params.append(int(checked))
    # 填报状态过滤（本周周一~周日区间）
    _current_round_for_filter = _current_fill_round()
    _flt_m, _flt_s = _week_range(_current_round_for_filter)
    if filled == '1':
        where_clauses.append(
            "id IN (SELECT hardware_id FROM key_account_followup_logs WHERE fill_round >= ? AND fill_round <= ?)"
        )
        params.extend([_flt_m, _flt_s])
    elif filled == '0':
        where_clauses.append(
            "id NOT IN (SELECT hardware_id FROM key_account_followup_logs WHERE fill_round >= ? AND fill_round <= ?)"
        )
        params.extend([_flt_m, _flt_s])
    if dept:
        where_clauses.append("department = ?")
        params.append(dept)
    if search:
        where_clauses.append("(customer_name LIKE ? OR department LIKE ? OR owner_name LIKE ?)")
        params += [f'%{search}%', f'%{search}%', f'%{search}%']

    where_sql = ('WHERE ' + ' AND '.join(where_clauses)) if where_clauses else ''

    total = db.execute(
        f"SELECT COUNT(*) FROM key_account_hardware {where_sql}", params
    ).fetchone()[0]

    # 全量合计（不分页，统计所有符合条件的记录）
    agg = db.execute(
        f"SELECT COALESCE(SUM(classroom_count),0), COALESCE(SUM(lent_units),0), "
        f"COALESCE(SUM(purchased_units),0), COALESCE(SUM(visit_count),0), "
        f"COALESCE(SUM(opportunity_count),0), COALESCE(SUM(opportunity_amount),0) "
        f"FROM key_account_hardware {where_sql}",
        params
    ).fetchone()

    current_round = _current_fill_round()
    cur_m, cur_s = _week_range(current_round)
    last_week_round = (dt.date.fromisoformat(current_round) - dt.timedelta(days=7)).isoformat()
    last_m, last_s = _week_range(last_week_round)

    # 本周已填报数（全量）
    base_where = (' AND '.join(where_clauses)) if where_clauses else '1=1'
    filled_count = db.execute(
        f"SELECT COUNT(*) FROM key_account_hardware h "
        f"WHERE h.id IN (SELECT hardware_id FROM key_account_followup_logs WHERE fill_round >= ? AND fill_round <= ?) "
        f"AND ({base_where})",
        [cur_m, cur_s] + params
    ).fetchone()[0]

    # 上周已填报数（全量）
    last_week_filled_count = db.execute(
        f"SELECT COUNT(*) FROM key_account_hardware h "
        f"WHERE h.id IN (SELECT hardware_id FROM key_account_followup_logs WHERE fill_round >= ? AND fill_round <= ?) "
        f"AND ({base_where})",
        [last_m, last_s] + params
    ).fetchone()[0]

    totals = {
        'classroom_count': agg[0],
        'lent_units': agg[1],
        'purchased_units': agg[2],
        'visit_count': agg[3],
        'opportunity_count': agg[4],
        'opportunity_amount': round(float(agg[5] or 0), 2),
        'last_week_filled': last_week_filled_count,
        'last_week_not_filled': total - last_week_filled_count,
        'this_week_filled': filled_count,
        'this_week_not_filled': total - filled_count,
        'total_records': total,
    }

    rows = db.execute(
        f"SELECT * FROM key_account_hardware {where_sql} ORDER BY classroom_count DESC LIMIT ? OFFSET ?",
        params + [page_size, (page - 1) * page_size]
    ).fetchall()

    items = [dict(r) for r in rows]

    # 附加最新填报信息
    for item in items:
        hw_id = item['id']
        # 最新一条填报
        latest = db.execute(
            "SELECT * FROM key_account_followup_logs WHERE hardware_id=? ORDER BY fill_round DESC LIMIT 1",
            (hw_id,)
        ).fetchone()
        if latest:
            item['latest_followup'] = dict(latest)
            item['latest_stage'] = latest['checkin_stage']
        else:
            item['latest_followup'] = None
            item['latest_stage'] = item.get('checkin_stage')

        # 本周是否已填报
        this_week = db.execute(
            "SELECT id, filled_at FROM key_account_followup_logs WHERE hardware_id=? AND fill_round>=? AND fill_round<=?",
            (hw_id, cur_m, cur_s)
        ).fetchone()
        item['this_week_filled'] = bool(this_week)

        # 上周是否已填报
        last_week = db.execute(
            "SELECT id FROM key_account_followup_logs WHERE hardware_id=? AND fill_round>=? AND fill_round<=?",
            (hw_id, last_m, last_s)
        ).fetchone()
        item['last_week_filled'] = bool(last_week)

    return jsonify({'total': total, 'page': page, 'page_size': page_size, 'items': items, 'totals': totals})


@app.route('/api/key-account/hardware/<int:record_id>', methods=['PUT'])
@token_required
def ka_hardware_update(record_id):
    """更新顾问手动填写字段"""
    data = request.get_json()
    db = get_db()

    manual_fields = ['is_checked_in', 'checkin_stage', 'stage_desc',
                     'feedback', 'has_blocker', 'blocker_detail', 'opportunity_status',
                     'stage4_reported', 'stage4_decision_role', 'stage4_meeting_held',
                     'stage4_meeting_files', 'stage6_no_intent_reason']
    updates = {k: data[k] for k in manual_fields if k in data}
    if not updates:
        return jsonify({'error': '无可更新字段'}), 400

    # 手动打点：若 is_checked_in 设为 1 且 checkin_date 为空，自动填充当前日期
    if updates.get('is_checked_in') == 1:
        existing = db.execute(
            "SELECT checkin_date FROM key_account_hardware WHERE id = ?", (record_id,)
        ).fetchone()
        if existing and (not existing['checkin_date'] or existing['checkin_date'] == ''):
            updates['checkin_date'] = dt.datetime.now().strftime('%Y-%m-%d')

    updates['updated_by'] = g.current_user.get('display_name', '')
    updates['updated_at'] = dt.datetime.now().isoformat()

    set_clause = ', '.join(f"{k} = ?" for k in updates)
    db.execute(
        f"UPDATE key_account_hardware SET {set_clause} WHERE id = ?",
        list(updates.values()) + [record_id]
    )
    db.commit()
    return jsonify({'message': '保存成功'})


@app.route('/api/key-account/hardware/<int:hw_id>/pause', methods=['PUT'])
@token_required
def ka_hardware_pause(hw_id):
    """暂停跟进某个客户"""
    data = request.get_json() or {}
    reason = (data.get('reason') or '').strip()[:200]
    if not reason:
        return jsonify({'error': '请选择暂停原因'}), 400
    db = get_db()
    row = db.execute("SELECT * FROM key_account_hardware WHERE id=?", (hw_id,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': '记录不存在'}), 404
    # 权限：管理员可操作任意，顾问只能操作自己的
    if g.current_user['role'] != 'admin':
        visible = get_visible_owners(db, g.current_user)
        if visible is not None:
            owner_name = row['owner_name'] or ''
            allowed = any(owner_name.startswith(f"{n}-") for n in visible)
            if not allowed:
                db.close()
                return jsonify({'error': '无权限操作此客户'}), 403
    db.execute("UPDATE key_account_hardware SET is_paused=1, pause_reason=?, paused_by=?, paused_at=? WHERE id=?",
               (reason, g.current_user['display_name'], dt.datetime.now().isoformat(), hw_id))
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': '已暂停跟进'})


@app.route('/api/key-account/hardware/<int:hw_id>/resume', methods=['PUT'])
@token_required
def ka_hardware_resume(hw_id):
    """恢复跟进某个客户"""
    db = get_db()
    row = db.execute("SELECT * FROM key_account_hardware WHERE id=?", (hw_id,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': '记录不存在'}), 404
    # 权限
    if g.current_user['role'] != 'admin':
        visible = get_visible_owners(db, g.current_user)
        if visible is not None:
            owner_name = row['owner_name'] or ''
            allowed = any(owner_name.startswith(f"{n}-") for n in visible)
            if not allowed:
                db.close()
                return jsonify({'error': '无权限操作此客户'}), 403
    db.execute("UPDATE key_account_hardware SET is_paused=0, pause_reason=NULL, paused_by=NULL, paused_at=NULL WHERE id=?",
               (hw_id,))
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': '已恢复跟进'})


@app.route('/api/key-account/hardware/sync', methods=['POST'])
@token_required
def ka_hardware_sync():
    """
    接收前端传入的 CRM 汇总数据并写入/更新本地缓存。
    前端负责调用 CRM MCP 聚合数据，后端只做 upsert。
    请求体：{ "items": [ { crm_account_id, customer_name, department, owner_name,
                            created_date, classroom_count,
                            lent_units, lent_date, purchased_units, purchase_date,
                            visit_count, last_visit_date, opportunity_count,
                            opportunity_amount, opportunity_status } ] }
    """
    data = request.get_json()
    items = data.get('items', [])
    if not items:
        return jsonify({'error': '数据为空'}), 400

    db = get_db()
    synced_at = dt.datetime.now().isoformat()
    upserted = 0
    for item in items:
        existing = db.execute(
            "SELECT id, lent_date, purchase_date FROM key_account_hardware WHERE crm_account_id = ?",
            (item.get('crm_account_id'),)
        ).fetchone()

        # 自动打点逻辑：借用或采购不为0 -> 默认已打点 + 阶段1
        lent = item.get('lent_units', 0) or 0
        purchased = item.get('purchased_units', 0) or 0
        auto_checked = 1 if (lent > 0 or purchased > 0) else 0
        auto_stage = '1' if auto_checked else None

        # 处理 lent_date / purchase_date：优先用传入值，否则保留已有值
        new_lent_date = item.get('lent_date', '') or ''
        new_purchase_date = item.get('purchase_date', '') or ''
        if existing:
            # 如果传入为空但已有值，保留已有值
            if not new_lent_date and existing['lent_date']:
                new_lent_date = existing['lent_date']
            if not new_purchase_date and existing['purchase_date']:
                new_purchase_date = existing['purchase_date']

        # 计算 checkin_date = MIN(lent_date, purchase_date)
        checkin_date = ''
        if new_lent_date and new_purchase_date:
            checkin_date = new_lent_date if new_lent_date < new_purchase_date else new_purchase_date
        elif new_lent_date:
            checkin_date = new_lent_date
        elif new_purchase_date:
            checkin_date = new_purchase_date

        if existing:
            # 不覆盖已有手动填写字段，但若未打点且满足自动条件则自动打点
            # 同时更新 lent_date/purchase_date/checkin_date（仅当传入非空值或保留已有值）
            db.execute('''UPDATE key_account_hardware
                SET customer_name=?, department=?, owner_name=?,
                    created_date=COALESCE(NULLIF(?, ''), created_date),
                    classroom_count=?, lent_units=?, purchased_units=?,
                    lent_date=CASE WHEN ? != '' THEN ? ELSE lent_date END,
                    purchase_date=CASE WHEN ? != '' THEN ? ELSE purchase_date END,
                    checkin_date=CASE WHEN ? != '' THEN ? ELSE checkin_date END,
                    visit_count=?, last_visit_date=?, opportunity_count=?,
                    opportunity_amount=?, opportunity_status=?,
                    is_checked_in = CASE WHEN is_checked_in=0 AND ?=1 THEN 1 ELSE is_checked_in END,
                    checkin_stage  = CASE WHEN is_checked_in=0 AND checkin_stage IS NULL AND ?=1 THEN '1' ELSE checkin_stage END,
                    synced_at=?
                WHERE crm_account_id=?''',
                (item.get('customer_name'), item.get('department'), item.get('owner_name'),
                 item.get('created_date', ''),
                 item.get('classroom_count', 0), lent, purchased,
                 new_lent_date, new_lent_date,
                 new_purchase_date, new_purchase_date,
                 checkin_date, checkin_date,
                 item.get('visit_count', 0), item.get('last_visit_date', ''),
                 item.get('opportunity_count', 0),
                 item.get('opportunity_amount', 0),
                 item.get('opportunity_status', ''),
                 auto_checked, auto_checked,
                 synced_at, item.get('crm_account_id')))
        else:
            db.execute('''INSERT INTO key_account_hardware
                (crm_account_id, customer_name, department, owner_name,
                 created_date, classroom_count,
                 lent_units, lent_date, purchased_units, purchase_date,
                 checkin_date,
                 visit_count, last_visit_date, opportunity_count,
                 opportunity_amount, opportunity_status,
                 is_checked_in, checkin_stage, synced_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (item.get('crm_account_id'), item.get('customer_name'),
                 item.get('department'), item.get('owner_name'),
                 item.get('created_date', ''),
                 item.get('classroom_count', 0),
                 lent, new_lent_date, purchased, new_purchase_date,
                 checkin_date,
                 item.get('visit_count', 0), item.get('last_visit_date', ''),
                 item.get('opportunity_count', 0),
                 item.get('opportunity_amount', 0),
                 item.get('opportunity_status', ''),
                 auto_checked, auto_stage,
                 synced_at))
        upserted += 1

    db.commit()
    return jsonify({'message': f'同步完成，共处理 {upserted} 条记录', 'synced_at': synced_at})


@app.route('/api/key-account/hardware/depts', methods=['GET'])
@token_required
def ka_hardware_depts():
    """返回客户清单中的部门列表（去重排序）"""
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT department FROM key_account_hardware WHERE department IS NOT NULL AND department != '' ORDER BY department"
    ).fetchall()
    return jsonify({'depts': [r[0] for r in rows]})


@app.route('/api/key-account/sync-crm-accounts', methods=['POST'])
@token_required
def ka_sync_crm_accounts():
    """
    代理前端从 CRM 拉取客户清单。
    当前版本返回本地数据库中已有的客户数据（格式兼容CRM字段），
    后续可接入 MCP entity.search 实现实时同步。
    请求体：{ page, page_size }
    """
    data = request.get_json() or {}
    page = int(data.get('page', 1))
    page_size = min(int(data.get('page_size', 500)), 500)
    offset = (page - 1) * page_size

    db = get_db()
    total = db.execute("SELECT COUNT(*) FROM key_account_hardware").fetchone()[0]
    rows = db.execute(
        "SELECT * FROM key_account_hardware ORDER BY classroom_count DESC LIMIT ? OFFSET ?",
        (page_size, offset)
    ).fetchall()

    records = []
    for r in rows:
        records.append({
            'id': r['crm_account_id'] or str(r['id']),
            'accountName': r['customer_name'],
            'name': r['customer_name'],
            'dimDepart': {'name': r['department'] or ''},
            'ownerId': {'name': r['owner_name'] or ''},
            'ClassroomN__c': r['classroom_count'] or 0,
            'newCreateDate__c': r['created_date'] or '',
            # 聚合字段暂不返回明细，由前端从本地DB读取
        })

    return jsonify({
        'records': records,
        'total': total,
        'page': page,
        'page_size': page_size,
    })


@app.route('/api/key-account/hardware/stats', methods=['GET'])
@token_required
def ka_hardware_stats():
    """返回统计卡片数据"""
    db = get_db()

    # 权限过滤
    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')
    base_where = f"WHERE {owner_filter}" if owner_filter else ""
    base_params = owner_params if owner_params else []

    # 默认排除暂停客户
    paused_cond = "COALESCE(is_paused,0)=0"
    base_where = f"{base_where} {'AND' if base_where else 'WHERE'} {paused_cond}"

    total = db.execute(
        f"SELECT COUNT(*) FROM key_account_hardware {base_where}", base_params
    ).fetchone()[0]

    checked_in = db.execute(
        f"SELECT COUNT(*) FROM key_account_hardware {base_where} {'AND' if base_where else 'WHERE'} is_checked_in=1",
        base_params
    ).fetchone()[0] if total else 0
    not_checked = total - checked_in

    # 阶段统计：优先取最新填报的阶段，无填报则取主表
    stage_counts = {}
    for i in range(1, 7):
        and_or = 'AND' if base_where else 'WHERE'
        # 有填报记录的客户 → 取最新填报阶段
        with_logs = db.execute(f"""
            SELECT COUNT(*) FROM (
                SELECT h.id FROM key_account_hardware h
                {base_where}
                JOIN key_account_followup_logs fl ON fl.hardware_id = h.id
                AND fl.id = (SELECT MAX(id) FROM key_account_followup_logs WHERE hardware_id = h.id)
                WHERE {'1=1 AND' if base_where else '1=1 AND'} fl.checkin_stage = ?
            )
        """, base_params + [str(i)]).fetchone()[0]
        # 无填报记录的客户 → 取主表阶段
        without_logs = db.execute(f"""
            SELECT COUNT(*) FROM key_account_hardware h
            {base_where}
            {and_or} h.checkin_stage = ?
            AND NOT EXISTS (SELECT 1 FROM key_account_followup_logs WHERE hardware_id = h.id)
        """, base_params + [str(i)]).fetchone()[0]
        stage_counts[str(i)] = with_logs + without_logs

    # 本周填报统计（周一~周日区间）
    current_round = _current_fill_round()
    sm_m, sm_s = _week_range(current_round)
    this_week_filled = db.execute(f"""
        SELECT COUNT(DISTINCT fl.hardware_id)
        FROM key_account_followup_logs fl
        JOIN key_account_hardware h ON h.id = fl.hardware_id
        {base_where}
        {'AND' if base_where else 'WHERE'} fl.fill_round >= ? AND fl.fill_round <= ?
    """, base_params + [sm_m, sm_s]).fetchone()[0]

    # 汇总借用/采购/拜访/商机
    sums = db.execute(f"""
        SELECT COALESCE(SUM(lent_units), 0), COALESCE(SUM(purchased_units), 0),
               COALESCE(SUM(visit_count), 0), COALESCE(SUM(opportunity_count), 0)
        FROM key_account_hardware {base_where}
    """, base_params).fetchone()

    return jsonify({
        'total': total,
        'checked_in': checked_in,
        'not_checked': not_checked,
        'stage_counts': stage_counts,
        'total_lent_units': sums[0],
        'total_purchased_units': sums[1],
        'total_visits': sums[2],
        'total_opportunities': sums[3],
        'this_week_filled': this_week_filled,
    })


# ============================================================
# 阶段配置 API（管理员可编辑）
# ============================================================

@app.route('/api/key-account/stages', methods=['GET'])
@token_required
def ka_stages_list():
    """获取阶段配置列表（所有人可读）"""
    db = get_db()
    rows = db.execute("SELECT * FROM stages_config ORDER BY sort_order").fetchall()
    return jsonify({'stages': [dict(r) for r in rows]})


@app.route('/api/key-account/stages', methods=['PUT'])
@token_required
@admin_required
def ka_stages_update():
    """管理员批量更新阶段配置"""
    data = request.get_json()
    stages = data.get('stages', [])
    if not stages:
        return jsonify({'error': '数据为空'}), 400

    db = get_db()
    for s in stages:
        db.execute(
            "UPDATE stages_config SET stage_name=?, stage_desc=? WHERE stage_key=?",
            (s.get('stage_name', ''), s.get('stage_desc', ''), s.get('stage_key'))
        )
    db.commit()
    return jsonify({'message': f'已更新 {len(stages)} 个阶段'})


# ============================================================
# 每周填报 API（followup_logs）
# ============================================================

LOCK_DAYS = 5  # 填报后5天锁定

def _current_fill_round():
    """计算当前填报轮次（当周周日日期）。
    Mon-Sun → 本周周日（整个自然周内均可填报本周）。
    """
    today = dt.date.today()
    weekday = today.weekday()  # Monday=0 ... Sunday=6
    # 本周日 = today + (6 - weekday)
    sunday = today + dt.timedelta(days=6 - weekday)
    return sunday.isoformat()

def _week_range(round_date_str):
    """给定周日日期字符串，返回该周的 (monday_str, sunday_str)。
    周一~周日为自然周，周一是周日-6天。
    """
    sunday = dt.date.fromisoformat(round_date_str)
    monday = sunday - dt.timedelta(days=6)
    return (monday.isoformat(), sunday.isoformat())

def _is_log_locked(filled_at_str, is_admin):
    """判断填报记录是否已锁定（距 filled_at 超过 LOCK_DAYS 天）。
    管理员不受限制。
    """
    if is_admin:
        return False
    if not filled_at_str:
        return False
    try:
        filled_dt = dt.datetime.fromisoformat(filled_at_str)
        age = (dt.datetime.now() - filled_dt).total_seconds() / 86400
        return age > LOCK_DAYS
    except (ValueError, TypeError):
        return False


@app.route('/api/key-account/followup/current-round', methods=['GET'])
@token_required
def followup_current_round():
    """返回当前填报轮次信息"""
    today = dt.date.today()
    current_round = _current_fill_round()
    # 填报窗口：当周一00:00 至 当周日 23:59
    round_date = dt.date.fromisoformat(current_round)
    # 轮次本身就是周日，即截止日
    sunday = round_date
    is_fillable = today <= sunday
    days_remaining = (sunday - today).days if is_fillable else 0

    return jsonify({
        'current_round': current_round,
        'is_fillable': is_fillable,
        'days_remaining': days_remaining,
        'lock_days': LOCK_DAYS,
        'today': today.isoformat(),
    })


@app.route('/api/key-account/hardware/<int:record_id>/followup-logs', methods=['GET'])
@token_required
def followup_logs_list(record_id):
    """获取某客户的所有填报历史"""
    db = get_db()
    is_admin = g.current_user.get('role') == 'admin'
    rows = db.execute(
        "SELECT * FROM key_account_followup_logs WHERE hardware_id=? ORDER BY fill_round DESC",
        (record_id,)
    ).fetchall()

    items = []
    for r in rows:
        d = dict(r)
        d['is_locked'] = _is_log_locked(d.get('filled_at'), is_admin)
        items.append(d)

    return jsonify({'items': items})


@app.route('/api/key-account/hardware/<int:record_id>/followup-logs', methods=['POST'])
@token_required
def followup_log_create(record_id):
    """新增（或更新）本周填报"""
    data = request.get_json()
    current_round = _current_fill_round()
    is_admin = g.current_user.get('role') == 'admin'
    db = get_db()

    # 检查本周是否已有填报（周一~周日区间内任意日期都算）
    fu_m, fu_s = _week_range(current_round)
    existing = db.execute(
        "SELECT * FROM key_account_followup_logs WHERE hardware_id=? AND fill_round>=? AND fill_round<=?",
        (record_id, fu_m, fu_s)
    ).fetchone()

    if existing:
        # 已存在 → 检查锁定
        if _is_log_locked(existing['filled_at'], is_admin):
            return jsonify({'error': '该填报已锁定，无法修改'}), 403
        # 更新
        log_fields = ['checkin_stage', 'stage_desc', 'feedback', 'has_blocker',
                      'blocker_detail', 'followup_plan', 'opportunity_status',
                      'stage4_reported', 'stage4_decision_role', 'stage4_meeting_held',
                      'stage4_meeting_files', 'stage6_no_intent_reason']
        updates = {k: data[k] for k in log_fields if k in data}
        updates['updated_at'] = dt.datetime.now().isoformat()
        set_clause = ', '.join(f"{k} = ?" for k in updates)
        db.execute(
            f"UPDATE key_account_followup_logs SET {set_clause} WHERE id=?",
            list(updates.values()) + [existing['id']]
        )
        db.commit()
        # 同步最新阶段到主表（保持主表 checkin_stage 与最近一次填报一致）
        if 'checkin_stage' in updates:
            db.execute(
                "UPDATE key_account_hardware SET checkin_stage=?, updated_at=datetime('now','localtime') WHERE id=?",
                (updates['checkin_stage'], record_id)
            )
            db.commit()
        return jsonify({'message': '本周填报已更新', 'id': existing['id']})
    else:
        # 新增
        now = dt.datetime.now().isoformat()
        db.execute('''INSERT INTO key_account_followup_logs
            (hardware_id, fill_round, fill_date, checkin_stage, stage_desc,
             feedback, has_blocker, blocker_detail, followup_plan, opportunity_status,
             stage4_reported, stage4_decision_role, stage4_meeting_held,
             stage4_meeting_files, stage6_no_intent_reason,
             filled_by, filled_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (record_id, current_round, now[:10],
             data.get('checkin_stage', ''), data.get('stage_desc', ''),
             data.get('feedback', ''), data.get('has_blocker', 0),
             data.get('blocker_detail', ''), data.get('followup_plan', ''),
             data.get('opportunity_status', ''),
             data.get('stage4_reported', 0), data.get('stage4_decision_role', ''),
             data.get('stage4_meeting_held', 0), data.get('stage4_meeting_files', ''),
             data.get('stage6_no_intent_reason', ''),
             g.current_user.get('display_name', ''), now, now))
        # 新增填报时：自动标记 is_checked_in=1，并同步阶段到主表
        db.execute('''UPDATE key_account_hardware SET
            is_checked_in = 1,
            checkin_stage = ?,
            checkin_date = CASE WHEN checkin_date IS NULL OR checkin_date = '' THEN ? ELSE checkin_date END,
            updated_at = datetime('now','localtime')
            WHERE id = ?''',
            (data.get('checkin_stage', ''), now[:10], record_id))
        db.commit()
        log_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        return jsonify({'message': '本周填报已提交', 'id': log_id})


@app.route('/api/key-account/hardware/<int:record_id>/followup-logs/<int:log_id>', methods=['PUT'])
@token_required
def followup_log_update(record_id, log_id):
    """更新指定填报（检查锁定状态）"""
    is_admin = g.current_user.get('role') == 'admin'
    db = get_db()
    log = db.execute(
        "SELECT * FROM key_account_followup_logs WHERE id=? AND hardware_id=?",
        (log_id, record_id)
    ).fetchone()
    if not log:
        return jsonify({'error': '填报记录不存在'}), 404

    if _is_log_locked(log['filled_at'], is_admin):
        return jsonify({'error': '该填报已锁定，无法修改'}), 403

    data = request.get_json()
    log_fields = ['checkin_stage', 'stage_desc', 'feedback', 'has_blocker',
                  'blocker_detail', 'followup_plan', 'opportunity_status',
                  'stage4_reported', 'stage4_decision_role', 'stage4_meeting_held',
                  'stage4_meeting_files', 'stage6_no_intent_reason']
    updates = {k: data[k] for k in log_fields if k in data}
    if not updates:
        return jsonify({'error': '无可更新字段'}), 400
    updates['updated_at'] = dt.datetime.now().isoformat()
    set_clause = ', '.join(f"{k} = ?" for k in updates)
    db.execute(
        f"UPDATE key_account_followup_logs SET {set_clause} WHERE id=?",
        list(updates.values()) + [log_id]
    )
    db.commit()
    return jsonify({'message': '填报已更新'})


@app.route('/api/key-account/hardware/<int:record_id>/followup-logs/<int:log_id>', methods=['DELETE'])
@token_required
def followup_log_delete(record_id, log_id):
    """删除指定填报记录（5天内 + 管理员不受限）"""
    is_admin = g.current_user.get('role') == 'admin'
    db = get_db()
    log = db.execute(
        "SELECT * FROM key_account_followup_logs WHERE id=? AND hardware_id=?",
        (log_id, record_id)
    ).fetchone()
    if not log:
        return jsonify({'error': '填报记录不存在'}), 404

    if _is_log_locked(log['filled_at'], is_admin):
        return jsonify({'error': '该填报已锁定，无法删除'}), 403

    db.execute("DELETE FROM key_account_followup_logs WHERE id=?", (log_id,))
    db.commit()
    return jsonify({'message': '填报记录已删除'})


# ============================================================
# 数据分析 API
# ============================================================

@app.route('/api/key-account/analytics/summary', methods=['GET'])
@token_required
def ka_analytics_summary():
    """KPI 汇总卡片数据"""
    db = get_db()

    base = "FROM key_account_hardware"
    w = []
    p = []
    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')
    if owner_filter:
        w.append(owner_filter)
        p.extend(owner_params)
    # 默认排除暂停客户
    w.append("COALESCE(is_paused, 0) = 0")
    wh = (" WHERE " + " AND ".join(w)) if w else ""

    # 本财年新增客户数：created_date >= 2026-04-01（客户清单客户）
    new_customers = db.execute(
        f"SELECT COUNT(*) {base} {wh} {'AND' if wh else 'WHERE'} created_date >= '2026-04-01'",
        p if wh else []
    ).fetchone()[0]

    # 累计打点客户数（全部，含新客户和老客户）
    checked = db.execute(f"SELECT COUNT(*) {base} {wh} {'AND' if wh else 'WHERE'} is_checked_in=1",
                         p if wh else []).fetchone()[0]

    # 新客户打点数（created_date >= 2026-04-01 且已打点）
    checked_new = db.execute(
        f"SELECT COUNT(*) {base} {wh} {'AND' if wh else 'WHERE'} is_checked_in=1 AND created_date >= '2026-04-01'",
        p if wh else []
    ).fetchone()[0]

    # 老客户打点数（created_date < 2026-04-01 且已打点）
    checked_old = db.execute(
        f"SELECT COUNT(*) {base} {wh} {'AND' if wh else 'WHERE'} is_checked_in=1 AND created_date < '2026-04-01'",
        p if wh else []
    ).fetchone()[0]

    # 借用转销售转化率：有借用且有采购的客户 / 有借用的客户
    if wh:
        lent_total = db.execute(f"SELECT COUNT(*) {base} {wh} AND lent_units>0", p).fetchone()[0]
        lent_to_sale = db.execute(f"SELECT COUNT(*) {base} {wh} AND lent_units>0 AND purchased_units>0", p).fetchone()[0]
    else:
        lent_total = db.execute(f"SELECT COUNT(*) {base} WHERE lent_units>0").fetchone()[0]
        lent_to_sale = db.execute(f"SELECT COUNT(*) {base} WHERE lent_units>0 AND purchased_units>0").fetchone()[0]
    conversion = round(lent_to_sale / lent_total * 100, 1) if lent_total > 0 else 0

    # 价值复盘完成率：优先取最新填报阶段，无填报则取主表
    # 有填报记录且阶段4/5/6
    with_logs_stage4plus = db.execute(f"""
        SELECT COUNT(*) FROM key_account_hardware h
        JOIN key_account_followup_logs fl ON fl.id = (
            SELECT MAX(id) FROM key_account_followup_logs WHERE hardware_id = h.id
        )
        {wh} {'AND' if wh else 'WHERE'} fl.checkin_stage IN ('4','5','6')
    """, p if wh else []).fetchone()[0]
    # 无填报记录但主表阶段4/5/6
    without_logs_stage4plus = db.execute(f"""
        SELECT COUNT(*) FROM key_account_hardware h
        {wh}
        {'AND' if wh else 'WHERE'} h.is_checked_in=1 AND h.checkin_stage IN ('4','5','6')
        AND NOT EXISTS (SELECT 1 FROM key_account_followup_logs WHERE hardware_id = h.id)
    """, p if wh else []).fetchone()[0]
    stage4plus = with_logs_stage4plus + without_logs_stage4plus
    review_rate = round(stage4plus / checked * 100, 1) if checked > 0 else 0

    # 本周填报数（周一~周日区间）
    current_round = _current_fill_round()
    ks_m, ks_s = _week_range(current_round)
    this_week_filled = db.execute(f"""
        SELECT COUNT(DISTINCT fl.hardware_id)
        FROM key_account_followup_logs fl
        JOIN key_account_hardware h ON h.id = fl.hardware_id
        {wh}
        {'AND' if wh else 'WHERE'} fl.fill_round >= ? AND fl.fill_round <= ?
    """, (p + [ks_m, ks_s]) if wh else [ks_m, ks_s]).fetchone()[0]

    return jsonify({
        'new_customers': new_customers,
        'checked_in': checked,
        'checked_new': checked_new,
        'checked_old': checked_old,
        'conversion': conversion,
        'review_rate': review_rate,
        'stage4plus': stage4plus,
        'this_week_filled': this_week_filled,
    })


@app.route('/api/key-account/analytics/new-customers-monthly', methods=['GET'])
@token_required
def ka_analytics_new_customers():
    """每月新增客户数 — 按部门分组（仅统计 created_date >= 2026-04-01）"""
    dept = request.args.get('dept', '')
    db = get_db()

    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')

    if dept:
        sql = """
            SELECT substr(created_date, 1, 7) as month,
                   COUNT(*) as cnt
            FROM key_account_hardware
            WHERE department = ? AND created_date >= '2026-04-01'"""
        params = [dept]
    else:
        sql = """
            SELECT department,
                   substr(created_date, 1, 7) as month,
                   COUNT(*) as cnt
            FROM key_account_hardware
            WHERE created_date >= '2026-04-01'"""
        params = []

    if owner_filter:
        sql += " AND " + owner_filter
        params.extend(owner_params)

    if dept:
        sql += " GROUP BY month ORDER BY month"
    else:
        sql += " GROUP BY department, month ORDER BY month, department"

    rows = db.execute(sql, params).fetchall()
    data = [dict(r) for r in rows if r['month'] and len(r['month']) == 7]
    return jsonify({'data': data})


@app.route('/api/key-account/analytics/cross-table', methods=['GET'])
@token_required
def ka_analytics_cross_table():
    """部门×月份交叉表（不受权限/筛选器影响，全量数据）
    返回：
      - months: ['2026-04', '2026-05', ...]  动态到当前月
      - depts: [部门列表]
      - new_customers: {dept: {month: cnt, __total__: N}}  新增客户数（created_date>=2026-04-01）
      - checkins_new: {dept: {month: cnt}}   新客户打点数
      - checkins_old: {dept: {month: cnt}}   老客户打点数（created_date<2026-04-01）
    """
    db = get_db()
    today = dt.date.today()

    # 生成 2026-04 到当前月的月份列表
    months = []
    cur = dt.date(2026, 4, 1)
    end = dt.date(today.year, today.month, 1)
    while cur <= end:
        months.append(cur.strftime('%Y-%m'))
        if cur.month == 12:
            cur = dt.date(cur.year + 1, 1, 1)
        else:
            cur = dt.date(cur.year, cur.month + 1, 1)

    # 1. 新增客户数：created_date >= '2026-04-01'，按 created_date 归月
    rows_nc = db.execute("""
        SELECT department,
               substr(created_date, 1, 7) as month,
               COUNT(*) as cnt
        FROM key_account_hardware
        WHERE created_date >= '2026-04-01' AND COALESCE(is_paused,0)=0
        GROUP BY department, month
    """).fetchall()
    new_customers = {}
    dept_total_nc = {}
    for r in rows_nc:
        dept = r['department'] or '未知'
        m = r['month'] or ''
        dept_total_nc[dept] = dept_total_nc.get(dept, 0) + r['cnt']
        if len(m) == 7:
            new_customers.setdefault(dept, {})[m] = r['cnt']
    for dept, total in dept_total_nc.items():
        new_customers.setdefault(dept, {})['__total__'] = total

    # 2. 新客户打点数：is_checked_in=1 且 created_date >= '2026-04-01'（归属日期=借用/采购/创建，不含填报当天）
    attr_expr = """COALESCE(
        CASE
          WHEN TRIM(lent_date) > '' AND TRIM(purchase_date) > '' THEN MIN(TRIM(lent_date), TRIM(purchase_date))
          WHEN TRIM(lent_date) > '' THEN TRIM(lent_date)
          WHEN TRIM(purchase_date) > '' THEN TRIM(purchase_date)
          ELSE TRIM(created_date)
        END,
        TRIM(created_date)
      )"""
    rows_cn = db.execute(f"""
        SELECT department,
               substr({attr_expr}, 1, 7) as month,
               COUNT(*) as cnt
        FROM key_account_hardware
        WHERE is_checked_in=1 AND created_date >= '2026-04-01' AND {attr_expr} >= '2026-04-01'
        GROUP BY department, month
    """).fetchall()
    checkins_new = {}
    for r in rows_cn:
        dept = r['department'] or '未知'
        m = r['month'] or ''
        if len(m) == 7:
            checkins_new.setdefault(dept, {})[m] = r['cnt']

    # 3. 老客户打点数：is_checked_in=1 且 created_date < '2026-04-01'（归属日期=借用/采购/创建，不含填报当天）
    rows_co = db.execute(f"""
        SELECT department,
               substr({attr_expr}, 1, 7) as month,
               COUNT(*) as cnt
        FROM key_account_hardware
        WHERE is_checked_in=1 AND created_date < '2026-04-01' AND {attr_expr} >= '2026-04-01'
        GROUP BY department, month
    """).fetchall()
    checkins_old = {}
    for r in rows_co:
        dept = r['department'] or '未知'
        m = r['month'] or ''
        if len(m) == 7:
            checkins_old.setdefault(dept, {})[m] = r['cnt']

    # 4. 各部门打点总数（新财年，归属日期=借用/采购/创建，与月份明细一致）
    rows_tc = db.execute(f"""
        SELECT department, COUNT(*) as cnt
        FROM key_account_hardware
        WHERE is_checked_in=1 AND {attr_expr} >= '2026-04-01'
        GROUP BY department
    """).fetchall()
    total_checkins = {}
    for r in rows_tc:
        dept = r['department'] or '未知'
        total_checkins[dept] = r['cnt']

    # 汇总所有部门
    all_depts = sorted(set(
        list(new_customers.keys()) +
        list(checkins_new.keys()) +
        list(checkins_old.keys()) +
        list(total_checkins.keys())
    ))

    return jsonify({
        'months': months,
        'depts': all_depts,
        'new_customers': new_customers,
        'checkins_new': checkins_new,
        'checkins_old': checkins_old,
        'total_checkins': total_checkins,
    })


@app.route('/api/key-account/analytics/new-checkins-monthly', methods=['GET'])
@token_required
def ka_analytics_new_checkins():
    """每月新增打点客户数。
    打点归属日期按优先级：1) 借用日期 lent_date  2) 采购日期 purchase_date
                         3) 新客户创建日期 created_date（不使用填报当天 checkin_date）。
    仅统计新财年 2026-04 起。
    """
    dept = request.args.get('dept', '')
    db = get_db()

    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')

    # 打点归属日期：借用日期 → 采购日期 → 新客户创建日期（取借用/采购中较早者）
    attr_expr = """COALESCE(
        CASE
          WHEN TRIM(lent_date) > '' AND TRIM(purchase_date) > '' THEN MIN(TRIM(lent_date), TRIM(purchase_date))
          WHEN TRIM(lent_date) > '' THEN TRIM(lent_date)
          WHEN TRIM(purchase_date) > '' THEN TRIM(purchase_date)
          ELSE TRIM(created_date)
        END,
        TRIM(created_date)
      )"""
    base_where = f"is_checked_in=1 AND {attr_expr} >= '2026-04-01'"
    if dept:
        sql = f"""
            SELECT substr({attr_expr}, 1, 7) as month,
                   COUNT(*) as cnt
            FROM key_account_hardware
            WHERE {base_where} AND department=?"""
        params = [dept]
    else:
        sql = f"""
            SELECT department,
                   substr({attr_expr}, 1, 7) as month,
                   COUNT(*) as cnt
            FROM key_account_hardware
            WHERE {base_where}"""
        params = []

    if owner_filter:
        sql += " AND " + owner_filter
        params.extend(owner_params)

    if dept:
        sql += " GROUP BY month ORDER BY month"
    else:
        sql += " GROUP BY department, month ORDER BY month, department"

    rows = db.execute(sql, params).fetchall()
    data = [dict(r) for r in rows if r['month'] and len(r['month']) == 7]
    return jsonify({'data': data})


@app.route('/api/key-account/analytics/fill-rounds', methods=['GET'])
@token_required
def ka_analytics_fill_rounds():
    """返回所有填报轮次列表（归一化到周日，用于筛选下拉）"""
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT date(fill_round, 'weekday 0') as fill_round FROM key_account_followup_logs ORDER BY fill_round DESC"
    ).fetchall()
    return jsonify({'rounds': [r['fill_round'] for r in rows]})


@app.route('/api/key-account/analytics/stage-distribution', methods=['GET'])
@token_required
def ka_analytics_stage_dist():
    """阶段分布占比。
    参数：
      - dept: 部门筛选
      - round: 指定周（fill_round），不传则取全部时间（优先最新填报）
    """
    dept = request.args.get('dept', '')
    round_param = request.args.get('round', '')
    db = get_db()

    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')

    if round_param:
        # 指定周：查该周周一~周日区间的 followup_logs
        sd_m, sd_s = _week_range(round_param)
        if dept:
            sql = """
                SELECT fl.checkin_stage, COUNT(DISTINCT fl.hardware_id) as cnt
                FROM key_account_followup_logs fl
                JOIN key_account_hardware h ON h.id = fl.hardware_id
                WHERE fl.fill_round >= ? AND fl.fill_round <= ? AND h.is_checked_in=1 AND COALESCE(h.is_paused,0)=0 AND h.department=?
            """
            params = [sd_m, sd_s, dept]
        else:
            sql = """
                SELECT h.department, fl.checkin_stage, COUNT(DISTINCT fl.hardware_id) as cnt
                FROM key_account_followup_logs fl
                JOIN key_account_hardware h ON h.id = fl.hardware_id
                WHERE fl.fill_round >= ? AND fl.fill_round <= ? AND h.is_checked_in=1 AND COALESCE(h.is_paused,0)=0
            """
            params = [sd_m, sd_s]

        if owner_filter:
            sql += " AND " + owner_filter.replace('owner_name', 'h.owner_name')
            params.extend(owner_params)

        if dept:
            sql += " GROUP BY fl.checkin_stage ORDER BY fl.checkin_stage"
        else:
            sql += " GROUP BY h.department, fl.checkin_stage ORDER BY h.department, fl.checkin_stage"
    else:
        # 全部时间：直接从主表取 checkin_stage（与阶段6明细保持一致）
        # 主表 checkin_stage 在每次填报时同步更新，是当前阶段的权威数据源
        if dept:
            sql = """
                SELECT h.checkin_stage as checkin_stage, COUNT(*) as cnt
                FROM key_account_hardware h
                WHERE h.is_checked_in=1 AND COALESCE(h.is_paused,0)=0 AND h.checkin_stage IS NOT NULL AND h.checkin_stage != '' AND h.department=?
            """
            params = [dept]
        else:
            sql = """
                SELECT h.department as department, h.checkin_stage as checkin_stage, COUNT(*) as cnt
                FROM key_account_hardware h
                WHERE h.is_checked_in=1 AND COALESCE(h.is_paused,0)=0 AND h.checkin_stage IS NOT NULL AND h.checkin_stage != ''
            """
            params = []

        if owner_filter:
            sql += " WHERE " + owner_filter
            params.extend(owner_params)

        if dept:
            sql += " GROUP BY checkin_stage ORDER BY checkin_stage"
        else:
            sql += " GROUP BY department, checkin_stage ORDER BY department, checkin_stage"

    rows = db.execute(sql, params).fetchall()
    return jsonify({'data': [dict(r) for r in rows]})


@app.route('/api/key-account/analytics/stage6-detail', methods=['GET'])
@token_required
def ka_analytics_stage6():
    """阶段6客户明细"""
    db = get_db()
    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'h.owner_name')

    sql = """
        SELECT h.customer_name, h.department, h.owner_name, h.classroom_count,
               h.lent_units, h.purchased_units,
               COALESCE(
                 (SELECT fl.stage6_no_intent_reason FROM key_account_followup_logs fl
                  WHERE fl.hardware_id = h.id AND fl.stage6_no_intent_reason IS NOT NULL AND fl.stage6_no_intent_reason != ''
                  ORDER BY fl.id DESC LIMIT 1),
                 h.stage6_no_intent_reason
               ) AS stage6_no_intent_reason,
               h.stage_desc, h.updated_at
        FROM key_account_hardware h
        WHERE h.checkin_stage = '6' AND COALESCE(h.is_paused,0)=0"""
    params = []

    if owner_filter:
        sql += " AND " + owner_filter
        params.extend(owner_params)

    sql += " ORDER BY h.classroom_count DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify({'items': [dict(r) for r in rows]})


@app.route('/api/key-account/analytics/weekly-fill-stats', methods=['GET'])
@token_required
def ka_analytics_weekly_fill_stats():
    """填报统计（按部门/城市）。
    参数：round=YYYY-MM-DD（不传则默认为本周）
    """
    db = get_db()
    round_param = request.args.get('round', '')
    if round_param == 'last':
        current_round = (dt.date.fromisoformat(_current_fill_round()) - dt.timedelta(days=7)).isoformat()
    elif round_param:
        current_round = round_param
    else:
        current_round = _current_fill_round()
    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'h.owner_name')

    wf_m, wf_s = _week_range(current_round)
    sql = f"""
        SELECT h.department,
               COUNT(DISTINCT h.id) as total,
               COUNT(DISTINCT CASE WHEN fl.id IS NOT NULL THEN h.id END) as filled,
               COUNT(DISTINCT CASE WHEN fl.id IS NULL THEN h.id END) as unfilled
        FROM key_account_hardware h
        LEFT JOIN key_account_followup_logs fl ON fl.hardware_id = h.id AND fl.fill_round >= ? AND fl.fill_round <= ?
        WHERE 1=1
    """
    params = [wf_m, wf_s]
    if owner_filter:
        sql += " AND " + owner_filter
        params.extend(owner_params)
    sql += " GROUP BY h.department ORDER BY h.department"

    rows = db.execute(sql, params).fetchall()
    return jsonify({'round': current_round, 'items': [dict(r) for r in rows]})


@app.route('/api/key-account/analytics/drill', methods=['GET'])
@token_required
def ka_analytics_drill():
    """下钻弹窗数据"""
    t = request.args.get('type', '')       # monthly_customers | monthly_checkins | stage
    dept = request.args.get('dept', '')
    month = request.args.get('month', '')   # 月份如 2026-04
    stage = request.args.get('stagev', '')  # 阶段值 1-6
    db = get_db()

    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')

    if t == 'monthly_customers' and month and dept:
        sql = """
            SELECT owner_name, COUNT(*) as cnt
            FROM key_account_hardware
            WHERE department=? AND created_date >= ?||'-01' AND created_date < date(?||'-01','+1 month')"""
        params = [dept, month, month]
        if owner_filter:
            sql += " AND " + owner_filter
            params.extend(owner_params)
        sql += " GROUP BY owner_name ORDER BY cnt DESC"
        rows = db.execute(sql, params).fetchall()
        return jsonify({'title': f'{dept} {month} 新增客户 | 顾问分布', 'items': [dict(r) for r in rows]})

    if t == 'monthly_checkins' and month and dept:
        sql = """
            SELECT owner_name, COUNT(*) as cnt
            FROM key_account_hardware
            WHERE department=? AND is_checked_in=1
              AND created_date >= ?||'-01' AND created_date < date(?||'-01','+1 month')"""
        params = [dept, month, month]
        if owner_filter:
            sql += " AND " + owner_filter
            params.extend(owner_params)
        sql += " GROUP BY owner_name ORDER BY cnt DESC"
        rows = db.execute(sql, params).fetchall()
        return jsonify({'title': f'{dept} {month} 新增打点 | 顾问分布', 'items': [dict(r) for r in rows]})

    if t == 'stage' and dept and stage:
        sql = """
            SELECT customer_name, owner_name, classroom_count,
                   lent_units, purchased_units, visit_count, opportunity_count,
                   opportunity_amount, stage_desc, feedback
            FROM key_account_hardware
            WHERE department=? AND checkin_stage=?"""
        params = [dept, stage]
        if owner_filter:
            sql += " AND " + owner_filter
            params.extend(owner_params)
        sql += " ORDER BY classroom_count DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify({'title': f'{dept} | 阶段{stage} 客户明细', 'items': [dict(r) for r in rows]})

    return jsonify({'error': '参数不足'}), 400


@app.route('/api/key-account/analytics/paused-list', methods=['GET'])
@token_required
def ka_analytics_paused_list():
    """暂停跟进客户清单（管理员看全部，普通用户看自己）"""
    db = get_db()
    reason = request.args.get('reason', '')
    dept = request.args.get('dept', '')
    search = request.args.get('q', '')
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))

    where = ["is_paused = 1"]
    params = []

    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')
    if owner_filter:
        where.append(owner_filter)
        params.extend(owner_params)

    if reason:
        where.append("pause_reason = ?")
        params.append(reason)
    if dept:
        where.append("department = ?")
        params.append(dept)
    if search:
        where.append("(customer_name LIKE ? OR department LIKE ? OR owner_name LIKE ?)")
        params += [f'%{search}%', f'%{search}%', f'%{search}%']

    where_sql = ' AND '.join(where)

    total = db.execute(f"SELECT COUNT(*) FROM key_account_hardware WHERE {where_sql}", params).fetchone()[0]

    rows = db.execute(f"""
        SELECT id, crm_account_id, customer_name, department, owner_name,
               classroom_count, lent_units, purchased_units,
               pause_reason, paused_by, paused_at, created_date
        FROM key_account_hardware
        WHERE {where_sql}
        ORDER BY paused_at DESC
        LIMIT ? OFFSET ?
    """, params + [page_size, (page - 1) * page_size]).fetchall()

    # 统计按原因分布
    reason_rows = db.execute(f"""
        SELECT pause_reason, COUNT(*) as cnt
        FROM key_account_hardware
        WHERE {where_sql.replace('pause_reason = ?', 'pause_reason IS NOT NULL')}
        GROUP BY pause_reason ORDER BY cnt DESC
    """, [p for p in params if reason or (reason and p == reason)] if reason else params).fetchall()

    return jsonify({
        'total': total,
        'items': [dict(r) for r in rows],
        'reason_distribution': [dict(r) for r in reason_rows],
        'page': page,
        'page_size': page_size,
    })


@app.route('/api/key-account/analytics/weekly-trend', methods=['GET'])
@token_required
def ka_analytics_weekly_trend():
    """按周看阶段变化趋势 — 每个填报轮次各阶段的客户数"""
    dept = request.args.get('dept', '')
    db = get_db()

    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'owner_name')

    sql = """
        SELECT date(fl.fill_round, 'weekday 0') as fill_round, fl.checkin_stage, COUNT(DISTINCT fl.hardware_id) as cnt,
               h.department as department
        FROM key_account_followup_logs fl
        JOIN key_account_hardware h ON h.id = fl.hardware_id
        WHERE fl.checkin_stage IS NOT NULL AND fl.checkin_stage != '' AND COALESCE(h.is_paused,0)=0
    """
    params = []

    if dept:
        sql += " AND h.department = ?"
        params.append(dept)

    if owner_filter:
        sql += " AND " + owner_filter.replace('owner_name', 'h.owner_name')
        params.extend(owner_params)

    sql += " GROUP BY date(fl.fill_round, 'weekday 0'), fl.checkin_stage ORDER BY date(fl.fill_round, 'weekday 0'), fl.checkin_stage"

    rows = db.execute(sql, params).fetchall()
    return jsonify({'data': [dict(r) for r in rows]})


@app.route('/api/key-account/hardware/owners', methods=['GET'])
@token_required
def ka_hardware_owners():
    """按部门筛选返回顾问列表（用于明细库筛选）"""
    dept = request.args.get('dept', '')
    db = get_db()
    sql = "SELECT DISTINCT owner_name FROM key_account_hardware WHERE owner_name IS NOT NULL AND owner_name != ''"
    params = []
    if dept:
        sql += " AND department = ?"
        params.append(dept)
    sql += " ORDER BY owner_name"
    rows = db.execute(sql, params).fetchall()
    owners = [r['owner_name'] for r in rows]
    return jsonify({'owners': owners})


@app.route('/api/key-account/followup-detail', methods=['GET'])
@token_required
def ka_followup_detail():
    """
    明细库矩阵数据：
    按月份返回所有客户的每周填报阶段，以及最新阶段和打点状态。
    参数：months=2026-06,2026-07（逗号分隔，支持多选）, dept=, owner=, checked=, q=
    """
    months_str = request.args.get('months', '')
    # 兼容旧的 month 参数
    if not months_str:
        months_str = request.args.get('month', '')
    dept = request.args.get('dept', '')
    owner = request.args.get('owner', '')
    checked = request.args.get('checked', '')   # ''=全部, '1'=已打点, '0'=未打点
    q = request.args.get('q', '').strip()

    if not months_str:
        return jsonify({'error': '请指定 months 参数，如 2026-06,2026-07'}), 400

    month_list = [m.strip() for m in months_str.split(',') if m.strip()]
    if not month_list:
        return jsonify({'error': 'months 参数为空'}), 400

    db = get_db()

    # 权限过滤
    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'h.owner_name')

    # 1. 计算所有选中月份的周日，合并去重
    import calendar
    month_names = {'01':'1月','02':'2月','03':'3月','04':'4月','05':'5月','06':'6月',
                   '07':'7月','08':'8月','09':'9月','10':'10月','11':'11月','12':'12月'}
    fridays = []        # 有序周日列表
    seen_sundays = set()

    for month in month_list:
        try:
            y, m = int(month[:4]), int(month[5:7])
        except:
            return jsonify({'error': f'month 格式错误：{month}'}), 400

        _, days_in_month = calendar.monthrange(y, m)
        for d in range(1, days_in_month + 1):
            dt_date = dt.date(y, m, d)
            if dt_date.weekday() == 6:   # Sunday
                iso = dt_date.isoformat()
                if iso not in seen_sundays:
                    seen_sundays.add(iso)
                    fridays.append(iso)

        # 该月无周日时兜底
        if not any(dt.date(y, m, d).weekday() == 6 for d in range(1, days_in_month + 1)):
            last_day = dt.date(y, m, days_in_month)
            sunday = last_day + dt.timedelta(days=(6 - last_day.weekday()) % 7)
            iso = sunday.isoformat()
            if iso not in seen_sundays:
                seen_sundays.add(iso)
                fridays.append(iso)

    fridays.sort()

    # 生成带月份的周标签
    week_labels = []
    month_week_counters = {}
    for friday in fridays:
        f_dt = dt.date.fromisoformat(friday)
        m_key = f_dt.strftime('%m')
        # 周一属于该周日-6天，用它判断周属于哪个月
        monday = f_dt - dt.timedelta(days=6)
        monday_month = monday.strftime('%m')
        # 默认用周一所在月份
        display_month_key = monday_month
        label_month_name = month_names.get(display_month_key, f'{int(display_month_key)}月')
        if display_month_key not in month_week_counters:
            month_week_counters[display_month_key] = 0
        month_week_counters[display_month_key] += 1
        w_num = month_week_counters[display_month_key]
        week_labels.append({'label': f'{label_month_name}W{w_num}', 'friday': friday})

    # 2. 构建客户基础查询
    where = "1=1"
    params = []

    if dept:
        where += " AND h.department = ?"
        params.append(dept)
    if owner:
        where += " AND h.owner_name = ?"
        params.append(owner)
    if checked == '1':
        where += " AND h.is_checked_in = 1"
    elif checked == '0':
        where += " AND (h.is_checked_in = 0 OR h.is_checked_in IS NULL)"
    if q:
        where += " AND (h.customer_name LIKE ? OR h.owner_name LIKE ?)"
        params.extend([f'%{q}%', f'%{q}%'])
    if owner_filter:
        where += " AND " + owner_filter
        params.extend(owner_params)

    # 3. 查询客户列表（含最新阶段 + 打点状态 + 业务字段）
    customers_sql = f"""
        SELECT h.id, h.customer_name, h.department, h.owner_name,
               h.is_checked_in, h.checkin_stage,
               COALESCE(h.classroom_count,0) as classroom_count,
               COALESCE(h.lent_units,0) as lent_units,
               COALESCE(h.purchased_units,0) as purchased_units,
               h.last_visit_date, h.checkin_date,
               (SELECT checkin_stage FROM key_account_followup_logs
                WHERE hardware_id = h.id ORDER BY fill_round DESC LIMIT 1) as latest_filled_stage
        FROM key_account_hardware h
        WHERE {where}
        ORDER BY h.department, h.owner_name, h.customer_name
    """
    customer_rows = db.execute(customers_sql, params).fetchall()
    customers = []
    hw_ids = []
    for r in customer_rows:
        d = dict(r)
        # 最新阶段：优先用最新填报的阶段，否则用主表阶段
        latest = d.get('latest_filled_stage') or d.get('checkin_stage') or ''
        d['latest_stage'] = latest
        d['weeks'] = {}
        customers.append(d)
        hw_ids.append(d['id'])

    if not hw_ids:
        return jsonify({'weeks': week_labels, 'customers': []})

    # 4. 批量查询这些客户在指定月份的填报记录（归一化 fill_round 到周日）
    placeholders = ','.join(['?'] * len(hw_ids))
    logs_sql = f"""
        SELECT hardware_id, date(fill_round, 'weekday 0') as fill_round, checkin_stage, has_blocker, blocker_detail,
               stage_desc, feedback, opportunity_status, filled_by, filled_at
        FROM key_account_followup_logs
        WHERE hardware_id IN ({placeholders}) AND date(fill_round, 'weekday 0') IN ({','.join(['?'] * len(fridays))})
        ORDER BY fill_round
    """
    # 精确匹配周五日期
    logs_params = list(hw_ids) + list(fridays)
    logs = db.execute(logs_sql, logs_params).fetchall()

    # 建立 hardware_id -> fill_round -> log 的映射
    log_map = {}
    for r in logs:
        hid = r['hardware_id']
        fr = r['fill_round']
        if hid not in log_map:
            log_map[hid] = {}
        log_map[hid][fr] = {
            'stage':              r['checkin_stage'] or '',
            'has_blocker':        r['has_blocker'],
            'blocker_detail':     r['blocker_detail'] or '',
            'stage_desc':         r['stage_desc'] or '',
            'feedback':           r['feedback'] or '',
            'opportunity_status': r['opportunity_status'] or '',
            'filled_by':          r['filled_by'] or '',
            'filled_at':          r['filled_at'] or '',
        }

    for c in customers:
        hid = c['id']
        if hid in log_map:
            c['weeks'] = log_map[hid]

    return jsonify({
        'weeks': week_labels,
        'customers': customers,
    })


# ============================================================
# 销售日报 API
# ============================================================

@app.route('/api/daily-report/summary', methods=['GET'])
@token_required
def daily_report_summary():
    """销售日报汇总数据。
    返回：
      - date_info: 今日/本周信息
      - today: 今日新增、今日打点、今日填报
      - this_week: 本周填报进度（总数/已填/未填/各部门）
      - recent: 最近填报记录（5条）
      - department_today: 今日各部门填报情况
    """
    db = get_db()
    today_str = dt.date.today().isoformat()

    owner_filter, owner_params = get_owner_filter(db, g.current_user, 'h.owner_name')

    # 本周信息
    current_round = _current_fill_round()
    wk_m, wk_s = _week_range(current_round)

    # 今日新增客户（created_date = 今天）
    new_cust_sql = "SELECT COUNT(*) FROM key_account_hardware WHERE created_date = ?"
    new_cust_params = [today_str]
    new_customers_today = db.execute(new_cust_sql, new_cust_params).fetchone()[0]

    # 今日新增打点（checkin_date = 今天）
    new_checkin_sql = "SELECT COUNT(*) FROM key_account_hardware WHERE is_checked_in=1 AND checkin_date = ?"
    new_checkin_today = db.execute(new_checkin_sql, [today_str]).fetchone()[0]

    # 今日填报数
    followup_today_sql = """
        SELECT COUNT(*) FROM key_account_followup_logs WHERE fill_date = ?
    """
    followup_today = db.execute(followup_today_sql, [today_str]).fetchone()[0]

    # 今日各部门填报数
    dept_today_sql = """
        SELECT h.department, COUNT(*) as cnt
        FROM key_account_followup_logs fl
        JOIN key_account_hardware h ON h.id = fl.hardware_id
        WHERE fl.fill_date = ?
    """
    dept_today_params = [today_str]
    if owner_filter:
        dept_today_sql += " AND " + owner_filter.replace('owner_name', 'h.owner_name')
        dept_today_params.extend(owner_params)
    dept_today_sql += " GROUP BY h.department ORDER BY cnt DESC"
    dept_today_rows = db.execute(dept_today_sql, dept_today_params).fetchall()
    department_today = [{'department': r['department'] or '未知', 'count': r['cnt']} for r in dept_today_rows]

    # 今日填报的详细记录（带客户名、阶段、顾问）
    followup_detail_sql = """
        SELECT fl.filled_by, fl.checkin_stage, fl.filled_at,
               h.customer_name, h.department
        FROM key_account_followup_logs fl
        JOIN key_account_hardware h ON h.id = fl.hardware_id
        WHERE fl.fill_date = ?
    """
    followup_detail_params = [today_str]
    if owner_filter:
        followup_detail_sql += " AND " + owner_filter.replace('owner_name', 'h.owner_name')
        followup_detail_params.extend(owner_params)
    followup_detail_sql += " ORDER BY fl.filled_at DESC LIMIT 10"
    followup_rows = db.execute(followup_detail_sql, followup_detail_params).fetchall()
    today_records = [{
        'customer_name': r['customer_name'],
        'department': r['department'],
        'stage': r['checkin_stage'] or '',
        'filled_by': r['filled_by'] or '',
        'filled_at': r['filled_at'] or '',
    } for r in followup_rows]

    # 本周填报进度（带权限过滤）
    where_sql = ""
    where_params = []
    if owner_filter:
        where_sql = " AND " + owner_filter.replace('owner_name', 'h.owner_name')
        where_params = list(owner_params)

    week_total_sql = f"SELECT COUNT(*) FROM key_account_hardware h WHERE 1=1{where_sql}"
    week_total = db.execute(week_total_sql, where_params).fetchone()[0]

    week_filled_sql = f"""
        SELECT COUNT(DISTINCT fl.hardware_id)
        FROM key_account_followup_logs fl
        JOIN key_account_hardware h ON h.id = fl.hardware_id
        WHERE fl.fill_round >= ? AND fl.fill_round <= ?{where_sql}
    """
    week_filled = db.execute(week_filled_sql, [wk_m, wk_s] + where_params).fetchone()[0]
    week_unfilled = max(0, week_total - week_filled)

    # 本周各部门填报进度
    week_dept_sql = f"""
        SELECT h.department,
               COUNT(DISTINCT h.id) as total,
               COUNT(DISTINCT CASE WHEN fl.id IS NOT NULL THEN h.id END) as filled
        FROM key_account_hardware h
        LEFT JOIN key_account_followup_logs fl ON fl.hardware_id = h.id AND fl.fill_round >= ? AND fl.fill_round <= ?
        WHERE 1=1{where_sql}
        GROUP BY h.department ORDER BY h.department
    """
    week_dept_rows = db.execute(week_dept_sql, [wk_m, wk_s] + where_params).fetchall()
    week_departments = [{
        'department': r['department'] or '未知',
        'total': r['total'],
        'filled': r['filled'],
        'unfilled': r['total'] - r['filled'],
        'rate': round(r['filled'] / r['total'] * 100, 1) if r['total'] > 0 else 0,
    } for r in week_dept_rows]

    # 最近5条填报记录（不限今日）
    recent_sql = """
        SELECT fl.filled_by, fl.checkin_stage, fl.filled_at,
               h.customer_name, h.department
        FROM key_account_followup_logs fl
        JOIN key_account_hardware h ON h.id = fl.hardware_id
        WHERE 1=1
    """
    recent_params = []
    if owner_filter:
        recent_sql += " AND " + owner_filter.replace('owner_name', 'h.owner_name')
        recent_params = list(owner_params)
    recent_sql += " ORDER BY fl.filled_at DESC LIMIT 5"
    recent_rows = db.execute(recent_sql, recent_params).fetchall()
    recent_records = [{
        'customer_name': r['customer_name'],
        'department': r['department'],
        'stage': r['checkin_stage'] or '',
        'filled_by': r['filled_by'] or '',
        'filled_at': r['filled_at'] or '',
    } for r in recent_rows]

    # 本周截止信息
    deadline_date = dt.date.fromisoformat(current_round)
    days_remaining = max(0, (deadline_date - dt.date.today()).days)

    return jsonify({
        'date_info': {
            'today': today_str,
            'current_round': current_round,
            'week_start': wk_m,
            'week_end': wk_s,
            'deadline_date': deadline_date.isoformat(),
            'days_remaining': days_remaining,
        },
        'today': {
            'new_customers': new_customers_today,
            'new_checkins': new_checkin_today,
            'followups': followup_today,
            'records': today_records,
            'department_breakdown': department_today,
        },
        'this_week': {
            'total': week_total,
            'filled': week_filled,
            'unfilled': week_unfilled,
            'rate': round(week_filled / week_total * 100, 1) if week_total > 0 else 0,
            'departments': week_departments,
        },
        'recent': recent_records,
    })


# ============================================================
# 静态文件
# ============================================================

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'login.html')

@app.route('/workbench')
def workbench():
    return send_from_directory(app.static_folder, 'workbench.html')

@app.route('/admin')
def admin_page():
    return send_from_directory(app.static_folder, 'admin.html')

@app.route('/key-account')
def key_account():
    return send_from_directory(app.static_folder, 'key_account.html')

@app.route('/key-account/hardware')
def hardware_followup():
    return send_from_directory(app.static_folder, 'hardware_followup.html')

@app.route('/key-account/analytics')
def hardware_analytics():
    return send_from_directory(app.static_folder, 'hardware_analytics.html')

@app.route('/key-account/followup-detail')
def followup_detail_page():
    return send_from_directory(app.static_folder, 'followup_detail.html')

@app.route('/key-account/data-import')
def data_import_page():
    return send_from_directory(app.static_folder, 'data_import.html')

@app.route('/daily-report')
def daily_report_page():
    return send_from_directory(app.static_folder, 'daily_report.html')

@app.route('/admin/permissions')
def permissions_page():
    return send_from_directory(app.static_folder, 'permissions.html')

# ============================================================
# 数据导入 API
# ============================================================

def _detect_data_type(filename):
    """根据文件名关键词自动识别数据类型（注意：具体类型优先，客户兜底）"""
    f = filename.lower()
    if '借用' in f:   return 'lent'
    if '采购' in f or '售卖' in f: return 'purchase'
    if '拜访' in f:   return 'visit'
    if '商机' in f:   return 'opportunity'
    if '客户' in f:   return 'customer'
    return 'customer'  # 默认


def _parse_file(file):
    """解析上传文件，返回 rows_data（CRM导出表头在第1行，数据从第2行开始）"""
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'csv'
    import tempfile, os as _os
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
    file.save(tmp.name)
    tmp.close()

    try:
        if ext == 'xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            ws = wb.active
            # CRM 导出的表头在第1行
            raw_headers = [cell.value for cell in ws[1]]
            # 裁剪尾部 None 列（openpyxl max_column 包含样式残留的空列）
            while raw_headers and raw_headers[-1] is None:
                raw_headers.pop()
            # 去重：重复列名只保留第一个
            seen = set()
            headers = []
            col_index = []  # 有效列索引（排除重复列）
            for i, h in enumerate(raw_headers):
                h_clean = str(h).strip() if h else ''
                if h_clean and h_clean not in seen:
                    seen.add(h_clean)
                    headers.append(h_clean)
                    col_index.append(i)
            rows_data = []
            for row_idx in range(2, ws.max_row + 1):
                vals = [ws.cell(row=row_idx, column=col_index[j]+1).value for j in range(len(headers))]
                # 至少有一个非空值才作为有效行
                if any(v is not None for v in vals):
                    rows_data.append(dict(zip(headers, vals)))
        else:
            import csv
            with open(tmp.name, 'r', encoding='utf-8-sig') as f:
                csv_reader = csv.reader(f)
                try:
                    raw_headers = next(csv_reader)
                except StopIteration:
                    rows_data = []
                else:
                    # 去重：重复列名加后缀 _dup，保留第一个
                    seen = set()
                    headers = []
                    col_index = []  # 有效列索引（排除重复列）
                    for i, h in enumerate(raw_headers):
                        h_clean = h.strip() if h else ''
                        if h_clean and h_clean not in seen:
                            seen.add(h_clean)
                            headers.append(h_clean)
                            col_index.append(i)
                    # 手动构建 dict，避免 DictReader 重复列覆盖
                    rows_data = []
                    for row in csv_reader:
                        if any(v and v.strip() for v in row):
                            d = {}
                            for j, idx in enumerate(col_index):
                                if idx < len(row):
                                    d[headers[j]] = row[idx]
                            rows_data.append(d)
    finally:
        try: _os.unlink(tmp.name)
        except: pass

    return rows_data


def _safe_crm_id(val):
    """安全转换CRM ID：处理Excel读取大数字时float精度问题（如 4.2867e+18 或 4286769953817391.0）"""
    if val is None:
        return ''
    if isinstance(val, float):
        # 大整数被Excel存为float，先转int再转str
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    # 去掉可能的 .0 后缀
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s

def _parse_purchase_pivot(file):
    """解析采购透视表（月度列展开的xlsx），返回 rows_data 列表，每行含 '客户','所属部门','客户所有人','ID','采购台数'"""
    import tempfile, openpyxl, os as _os
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    file.save(tmp.name)
    tmp.close()
    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb.active
        rows = []
        # 透视表结构：Rows 1-3 表头, Row 4 合计, Row 5+ 数据
        # 数据行: col0=部门, col1=所有人, col2=客户名, col3=ID, col4+=月度数值
        for row_idx in range(5, ws.max_row + 1):
            dept = ws.cell(row=row_idx, column=1).value
            owner = ws.cell(row=row_idx, column=2).value
            name = ws.cell(row=row_idx, column=3).value
            crm_id = ws.cell(row=row_idx, column=4).value
            if not name or str(name).strip() == '':
                continue
            if str(name).strip() == '合计':
                continue
            total = 0
            for col in range(5, ws.max_column + 1):
                v = ws.cell(row=row_idx, column=col).value
                if v is not None:
                    try:
                        fv = float(str(v).replace(',', ''))
                        if fv > 0:
                            total += fv
                    except (ValueError, TypeError):
                        pass
            rows.append({
                '客户': str(name).strip(),
                '所属部门': str(dept).strip() if dept else '',
                '客户所有人': str(owner).strip() if owner else '',
                'ID': _safe_crm_id(crm_id),
                '采购台数': int(total)
            })
        return rows
    finally:
        try: _os.unlink(tmp.name)
        except: pass


def _get_col(row, *keys, default=None):
    """从数据行中按优先级获取列值（兼容CRM导出列名和内部key）"""
    for k in keys:
        v = row.get(k)
        if v is not None:
            return v
    return default


# CRM导出列名 → 内部key 映射（用于客户名称统一匹配）
CUSTOMER_NAME_KEYS = ('客户', '客户名称', 'customer_name')
CUSTOMER_ID_KEYS = ('ID', '客户ID', 'crm_account_id')
CUSTOMER_OWNER_KEYS = ('客户所有人', '客户顾问', 'owner_name', '所有人')
CUSTOMER_DEPT_KEYS = ('所属部门', 'department')
CUSTOMER_CLASSROOM_KEYS = ('教室数量', '教室数', 'classroom_count')
CUSTOMER_CREATED_KEYS = ('日期', '首次合作日期', '创建日期', '客户创建日期', 'created_date')
CUSTOMER_SALES_COUNT_KEYS = ('教学一体机售卖台数', '一体机售卖台数', '售卖台数', 'sales_count')
LENT_DATE_KEYS = ('订单日期', '借用日期', 'lent_date')
PURCHASE_DATE_KEYS = ('日期', '采购日期', 'purchase_date')


@app.route('/api/data-import/upload', methods=['POST'])
@token_required
@admin_required
def data_import_upload():
    """上传并导入数据文件 — 支持多文件批量上传、文件夹上传、自动识别类型"""
    import json as _json

    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '请选择文件'}), 400

    # 三步处理确保顺序：客户清单 → 借用明细 → 其他（采购/拜访/商机）
    # 借用可能补充新客户，必须在客户清单之后但在其他之前
    customer_files = [f for f in files if _detect_data_type(f.filename) == 'customer']
    lent_files = [f for f in files if _detect_data_type(f.filename) == 'lent']
    other_files = [f for f in files if _detect_data_type(f.filename) not in ('customer', 'lent')]
    sorted_files = customer_files + lent_files + other_files
    print(f"[import order] customer={len(customer_files)}, lent={len(lent_files)}, other={len(other_files)}")

    mode = request.form.get('mode', 'replace')  # replace | upsert
    db = get_db()
    username = g.current_user.get('display_name', '')

    all_results = []
    total_success = 0
    total_errors = 0

    # 拜访数据跨文件累积（多个拜访文件合并处理）
    visit_accumulated = []  # [(crm_id, name, date_str), ...]

    for file in sorted_files:
        if file.filename == '':
            continue

        # 自动识别数据类型（文件名优先，其次用表单参数）
        data_type = _detect_data_type(file.filename)

        try:
            # 采购透视表（xlsx）需要特殊解析，扁平CSV用通用解析
            if data_type == 'purchase' and file.filename.lower().endswith('.xlsx'):
                rows_data = _parse_purchase_pivot(file)
            else:
                rows_data = _parse_file(file)
            total_rows = len(rows_data)
            success = 0
            errors = 0
            filtered = 0

            if data_type == 'customer':
                # 【重要】不再 DELETE，改为 UPSERT 保留前端手动填报数据
                # replace模式仅清零自动同步字段（后续 import 会重新填充）
                if mode == 'replace':
                    db.execute("UPDATE key_account_hardware SET lent_units=0, purchased_units=0, visit_count=0, last_visit_date='', opportunity_count=0, opportunity_amount=0, lent_date='', purchase_date='', checkin_date=''")

                for row in rows_data:
                    try:
                        crm_id = str(_get_col(row, *CUSTOMER_ID_KEYS) or '')
                        name = str(_get_col(row, *CUSTOMER_NAME_KEYS) or '')
                        owner = str(_get_col(row, *CUSTOMER_OWNER_KEYS) or '')
                        dept = str(_get_col(row, *CUSTOMER_DEPT_KEYS) or '')
                        classroom = int(_get_col(row, *CUSTOMER_CLASSROOM_KEYS) or 0)
                        created = str(_get_col(row, *CUSTOMER_CREATED_KEYS) or '')
                        created = created.replace('/', '-') if created else ''

                        # 过滤：教学一体机售卖台数 >= 5 的不统计
                        sales_str = str(_get_col(row, *CUSTOMER_SALES_COUNT_KEYS) or '')
                        sales_count = 0
                        try:
                            sales_count = int(float(sales_str.replace(',', '').replace('--', '0')))
                        except (ValueError, TypeError):
                            sales_count = 0
                        if sales_count >= 5:
                            filtered += 1
                            continue

                        if not name:
                            errors += 1
                            continue

                        existing = db.execute(
                            "SELECT id FROM key_account_hardware WHERE crm_account_id = ?",
                            (crm_id,)
                        ).fetchone() if crm_id else None

                        if existing:
                            # 仅更新 CRM 来源字段，保留手动填报字段不变
                            db.execute("""
                                UPDATE key_account_hardware SET
                                    customer_name=?, department=?, owner_name=?,
                                    classroom_count=?, created_date=?,
                                    updated_at=datetime('now','localtime')
                                WHERE id=?
                            """, (name, dept, owner, classroom, created, existing['id']))
                        else:
                            db.execute("""
                                INSERT INTO key_account_hardware
                                    (crm_account_id, customer_name, department, owner_name,
                                     classroom_count, created_date, updated_at)
                                VALUES (?, ?, ?, ?, ?, ?, datetime('now','localtime'))
                            """, (crm_id, name, dept, owner, classroom, created))
                        success += 1
                    except Exception as e:
                        errors += 1
                        print(f"Error importing customer row: {e}")

            elif data_type == 'lent':
                # 借用明细：按 crm_account_id 聚合（同一客户可能多行），再匹配更新
                # 新逻辑：借用中不重复的客户自动补充到客户清单基础表
                if mode == 'replace':
                    db.execute("UPDATE key_account_hardware SET lent_units = 0, lent_date = ''")

                # 第一步：借用中不重复的客户自动补充到客户清单（含 created_date 和 lent_date）
                new_cust = 0
                seen_ids = set()
                for row in rows_data:
                    try:
                        crm_id = str(_get_col(row, *CUSTOMER_ID_KEYS, default='') or '')
                        if not crm_id or crm_id in seen_ids:
                            continue
                        seen_ids.add(crm_id)
                        existing = db.execute(
                            "SELECT id FROM key_account_hardware WHERE crm_account_id = ?",
                            (crm_id,)
                        ).fetchone()
                        if not existing:
                            name = str(_get_col(row, *CUSTOMER_NAME_KEYS) or '')
                            dept = str(_get_col(row, *CUSTOMER_DEPT_KEYS) or '')
                            owner = str(_get_col(row, *CUSTOMER_OWNER_KEYS) or '')
                            classroom_str = str(_get_col(row, *CUSTOMER_CLASSROOM_KEYS, default='0') or '0')
                            try:
                                classroom = int(float(classroom_str.replace(',', '').replace('--', '0')))
                            except (ValueError, TypeError):
                                classroom = 0
                            # 采集客户创建日期和借用日期
                            created = str(_get_col(row, *CUSTOMER_CREATED_KEYS) or '')
                            created = created.replace('/', '-') if created else ''
                            lent_date = str(_get_col(row, *LENT_DATE_KEYS) or '')
                            lent_date = lent_date.replace('/', '-') if lent_date else ''
                            if name:
                                db.execute(
                                    "INSERT INTO key_account_hardware (crm_account_id, customer_name, department, owner_name, classroom_count, created_date, lent_date, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))",
                                    (crm_id, name, dept, owner, classroom, created, lent_date))
                                new_cust += 1
                    except Exception as e:
                        print(f"Error checking new customer from lent: {e}")

                # 第二步：聚合借用数量和日期
                lent_agg = {}  # {crm_account_id: {'name': ..., 'units': int, 'lent_date': str}}
                for row in rows_data:
                    try:
                        crm_id = str(_get_col(row, *CUSTOMER_ID_KEYS, default='') or '')
                        name = str(_get_col(row, *CUSTOMER_NAME_KEYS) or '')
                        units = int(float(str(_get_col(row, '一体机借用数量', '借用台数', 'lent_units', default=0) or 0).replace(',', '')) or 0)
                        date_str = str(_get_col(row, *LENT_DATE_KEYS) or '')
                        date_str = date_str.replace('/', '-') if date_str else ''
                        if not crm_id and not name:
                            continue
                        key = crm_id or name
                        if key not in lent_agg:
                            lent_agg[key] = {'crm_id': crm_id, 'name': name, 'units': 0, 'lent_date': ''}
                        lent_agg[key]['units'] += units
                        # 取最早的借用日期
                        if date_str and (not lent_agg[key]['lent_date'] or date_str < lent_agg[key]['lent_date']):
                            lent_agg[key]['lent_date'] = date_str
                    except Exception as e:
                        errors += 1
                        print(f"Error parsing lent row: {e}")

                # 第三步：匹配并更新
                matched = 0
                for key, info in lent_agg.items():
                    try:
                        existing = None
                        if info['crm_id']:
                            existing = db.execute(
                                "SELECT id FROM key_account_hardware WHERE crm_account_id = ?",
                                (info['crm_id'],)
                            ).fetchone()
                        if not existing and info['name']:
                            existing = db.execute(
                                "SELECT id FROM key_account_hardware WHERE customer_name = ?",
                                (info['name'],)
                            ).fetchone()
                        if not existing and info['name']:
                            existing = db.execute(
                                "SELECT id FROM key_account_hardware WHERE customer_name LIKE ?",
                                (f"%{info['name']}%",)
                            ).fetchone()
                        if existing:
                            db.execute(
                                "UPDATE key_account_hardware SET lent_units = ?, lent_date = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                                (info['units'], info['lent_date'] or None, existing['id'])
                            )
                            success += 1
                            matched += 1
                    except Exception as e:
                        errors += 1
                        print(f"Error updating lent for {key}: {e}")
                print(f"[lent import] total_rows={total_rows}, agg_customers={len(lent_agg)}, matched={matched}, new_cust={new_cust}")

            elif data_type == 'purchase':
                # 采购明细：按 crm_account_id/客户名称匹配，更新 purchased_units 和 purchase_date
                if mode == 'replace':
                    db.execute("UPDATE key_account_hardware SET purchased_units = 0, purchase_date = ''")

                matched = 0
                for row in rows_data:
                    try:
                        crm_id = _safe_crm_id(row.get('ID', '') or '')
                        name = str(row.get('客户', '') or '')
                        units = int(float(str(row.get('一体机采购数量', row.get('采购台数', 0)) or 0).replace(',', '')) or 0)
                        date_str = str(_get_col(row, *PURCHASE_DATE_KEYS) or '')
                        date_str = date_str.replace('/', '-') if date_str else ''
                        if not name and not crm_id:
                            errors += 1
                            continue

                        existing = None
                        # 优先用 crm_account_id 精确匹配
                        if crm_id:
                            existing = db.execute(
                                "SELECT id FROM key_account_hardware WHERE crm_account_id = ?",
                                (crm_id,)
                            ).fetchone()
                        # 其次按客户名称精确匹配
                        if not existing and name:
                            existing = db.execute(
                                "SELECT id FROM key_account_hardware WHERE customer_name = ?",
                                (name,)
                            ).fetchone()
                        # 兜底：按客户名称模糊匹配
                        if not existing and name:
                            existing = db.execute(
                                "SELECT id FROM key_account_hardware WHERE customer_name LIKE ?",
                                (f'%{name}%',)
                            ).fetchone()
                        if existing:
                            db.execute(
                                "UPDATE key_account_hardware SET purchased_units = ?, purchase_date = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                                (units, date_str or None, existing['id'])
                            )
                            success += 1
                            matched += 1
                    except Exception as e:
                        errors += 1
                        print(f"Error importing purchase row: {e}")
                # 输出未匹配的采购行（方便排查）
                unmatched = total_rows - matched - errors if total_rows else 0
                print(f"[purchase import] total_rows={total_rows}, matched={matched}, errors={errors}")

            elif data_type == 'visit':
                # 拜访明细：跨文件累积原始行，等全部文件处理后统一聚合匹配
                # 新拜访文件列: 客户, ID, 日期
                for row in rows_data:
                    try:
                        crm_id = str(_get_col(row, 'ID', *CUSTOMER_ID_KEYS, default='') or '')
                        name = str(_get_col(row, *CUSTOMER_NAME_KEYS) or '')
                        date_str = str(_get_col(row, '日期', '创建日期', '最近拜访日期', 'last_visit_date', default='') or '')
                        date_str = date_str.replace('/', '-') if date_str else ''
                        if crm_id or name:
                            visit_accumulated.append((crm_id, name, date_str))
                    except Exception:
                        pass
                # 记录文件级日志但不更新DB（统一在后面处理）
                db.execute("""
                    INSERT INTO data_import_log (file_name, data_type, row_count, success_count, error_count, imported_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (file.filename, data_type, total_rows, total_rows, 0, username))
                all_results.append({
                    'file': file.filename,
                    'data_type': data_type,
                    'total_rows': total_rows,
                    'success_count': total_rows,
                    'error_count': 0,
                })
                total_success += total_rows
                continue  # 跳过后续统一日志写入

            elif data_type == 'opportunity':
                # 商机明细：按 crm_account_id 匹配，聚合商机数和金额
                # CRM列名: ID → crm_account_id, 商机 → count, 商机金额 → amount, 商机进度 → status
                # 聚合：同一客户的商机记录求和（数量、金额），商机阶段取最晚日期的商机进度
                # 注意：商机不自动补充客户，只更新已在基础表中的客户
                if mode == 'replace':
                    db.execute("UPDATE key_account_hardware SET opportunity_count = 0, opportunity_amount = 0, opportunity_status = ''")

                # 聚合商机数据
                opp_agg = {}  # {crm_id: [count, amount, latest_date_str, stage]}
                for row in rows_data:
                    try:
                        crm_id = str(_get_col(row, 'ID', *CUSTOMER_ID_KEYS, default='') or '')
                        count = int(float(str(_get_col(row, '商机', '商机数', 'opportunity_count', default=0) or 0).replace(',', '')) or 0)
                        amount = float(str(_get_col(row, '商机金额', 'opportunity_amount', default=0) or 0).replace(',', '') or 0)
                        # 商机阶段：取"商机进度"列
                        stage = str(_get_col(row, '商机进度', '商机阶段', 'opportunity_status', default='') or '')
                        date_str = str(_get_col(row, '日期', '创建日期', 'opp_date', default='') or '')
                        date_str = date_str.replace('/', '-') if date_str else ''
                        if crm_id:
                            if crm_id not in opp_agg:
                                opp_agg[crm_id] = [0, 0.0, '', '']
                            opp_agg[crm_id][0] += count
                            opp_agg[crm_id][1] += amount
                            # 取最晚日期的商机进度
                            if date_str and (not opp_agg[crm_id][2] or date_str > opp_agg[crm_id][2]):
                                opp_agg[crm_id][2] = date_str
                                opp_agg[crm_id][3] = stage
                    except Exception as e:
                        errors += 1

                # 第三步：匹配并更新
                matched = 0
                for crm_id, (count, amount, _, stage) in opp_agg.items():
                    try:
                        existing = db.execute(
                            "SELECT id FROM key_account_hardware WHERE crm_account_id = ?",
                            (crm_id,)
                        ).fetchone()
                        if existing:
                            db.execute(
                                "UPDATE key_account_hardware SET opportunity_count = ?, opportunity_amount = ?, opportunity_status = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                                (count, amount, stage or None, existing['id'])
                            )
                            success += 1
                            matched += 1
                    except Exception as e:
                        errors += 1
                        print(f"Error importing opportunity row: {e}")
                print(f"[opportunity import] total_rows={total_rows}, agg_customers={len(opp_agg)}, matched={matched}")

            # 写入日志
            db.execute("""
                INSERT INTO data_import_log (file_name, data_type, row_count, success_count, error_count, filtered_count, imported_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (file.filename, data_type, total_rows, success, errors, filtered, username))

            result_entry = {
                'file': file.filename,
                'data_type': data_type,
                'total_rows': total_rows,
                'success_count': success,
                'error_count': errors,
            }
            if filtered > 0:
                result_entry['filtered_count'] = filtered
                result_entry['filtered_reason'] = '教学一体机售卖台数≥5'
            all_results.append(result_entry)
            total_success += success
            total_errors += errors

        except Exception as e:
            all_results.append({
                'file': file.filename,
                'data_type': data_type,
                'error': str(e),
            })
            total_errors += 1
            print(f"Error importing file {file.filename}: {e}")

    # === 拜访数据统一处理（跨文件累积后一次性聚合+匹配） ===
    if visit_accumulated:
        # 按 crm_account_id 聚合：COUNT + MAX日期
        visit_agg = {}  # {crm_id or name: {'crm_id': ..., 'name': ..., 'count': int, 'last_date': str}}
        for crm_id, name, date_str in visit_accumulated:
            key = crm_id or name
            if not key:
                continue
            if key not in visit_agg:
                visit_agg[key] = {'crm_id': crm_id, 'name': name, 'count': 0, 'last_date': ''}
            visit_agg[key]['count'] += 1
            if date_str and (not visit_agg[key]['last_date'] or date_str > visit_agg[key]['last_date']):
                visit_agg[key]['last_date'] = date_str

        visit_matched = 0
        for key, info in visit_agg.items():
            try:
                existing = None
                if info['crm_id']:
                    existing = db.execute(
                        "SELECT id FROM key_account_hardware WHERE crm_account_id = ?",
                        (info['crm_id'],)
                    ).fetchone()
                if not existing and info['name']:
                    existing = db.execute(
                        "SELECT id FROM key_account_hardware WHERE customer_name = ?",
                        (info['name'],)
                    ).fetchone()
                if not existing and info['name']:
                    existing = db.execute(
                        "SELECT id FROM key_account_hardware WHERE customer_name LIKE ?",
                        (f"%{info['name']}%",)
                    ).fetchone()
                if existing:
                    db.execute(
                        "UPDATE key_account_hardware SET visit_count = ?, last_visit_date = ?, updated_at = datetime('now','localtime') WHERE id = ?",
                        (info['count'], info['last_date'], existing['id'])
                    )
                    visit_matched += 1
            except Exception as e:
                print(f"Error updating visit for {key}: {e}")
        print(f"[visit combined] total_raw_rows={len(visit_accumulated)}, agg_customers={len(visit_agg)}, matched={visit_matched}")

    # === 计算打点日期 (checkin_date = MIN(lent_date, purchase_date)) ===
    db.execute("""
        UPDATE key_account_hardware SET checkin_date =
            CASE
                WHEN lent_date != '' AND purchase_date != '' AND lent_date < purchase_date THEN lent_date
                WHEN lent_date != '' AND purchase_date != '' AND purchase_date <= lent_date THEN purchase_date
                WHEN lent_date != '' AND purchase_date = '' THEN lent_date
                WHEN lent_date = '' AND purchase_date != '' THEN purchase_date
                ELSE checkin_date
            END
        WHERE lent_date != '' OR purchase_date != ''
    """)

    # 自动打点：借用或采购不为0 → 自动标记已打点 + 阶段1
    # 仅对未手动打点的客户生效（is_checked_in=0）
    db.execute("UPDATE key_account_hardware SET is_checked_in=1, checkin_stage='1', updated_at=datetime('now','localtime') WHERE is_checked_in=0 AND (lent_units > 0 OR purchased_units > 0)")

    # 取消自动打点：借用和采购都为0 且 阶段仍为1（非手动变更）→ 重置为未打点
    db.execute("UPDATE key_account_hardware SET is_checked_in=0, checkin_stage='', checkin_date='', updated_at=datetime('now','localtime') WHERE is_checked_in=1 AND checkin_stage='1' AND lent_units=0 AND purchased_units=0")

    db.commit()

    return jsonify({
        'success': True,
        'total_success': total_success,
        'total_errors': total_errors,
        'files': all_results,
    })


@app.route('/api/data-import/logs', methods=['GET'])
@token_required
@admin_required
def data_import_logs():
    """导入日志"""
    db = get_db()
    rows = db.execute(
        "SELECT * FROM data_import_log ORDER BY imported_at DESC LIMIT 20"
    ).fetchall()
    return jsonify({'logs': [dict(r) for r in rows]})


# ============================================================
# 权限管理 API
# ============================================================

@app.route('/api/admin/users', methods=['GET'])
@token_required
@admin_required
def admin_list_users():
    """管理员查看所有用户"""
    db = get_db()
    rows = db.execute("""
        SELECT id, username, display_name, role, roles, job_number,
               dept_l3, dept_l4, l3_leader, l4_leader, is_l3_leader, is_l4_leader,
               grade, job_title, enabled, email
        FROM users ORDER BY username
    """).fetchall()
    return jsonify({'users': [dict(r) for r in rows]})


@app.route('/api/admin/users', methods=['POST'])
@token_required
@admin_required
def admin_create_user():
    """管理员新增用户"""
    data = request.get_json()
    db = get_db()

    username = (data.get('username') or '').strip()
    if not username:
        return jsonify({'error': '登录名不能为空'}), 400

    # Check duplicate
    existing = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        return jsonify({'error': f'用户 {username} 已存在'}), 409

    display_name = data.get('display_name', username)
    role = data.get('role', 'user')
    roles = data.get('roles', '')
    dept_l3 = data.get('dept_l3') or None
    dept_l4 = data.get('dept_l4') or None
    is_l3_leader = int(data.get('is_l3_leader', 0))
    is_l4_leader = int(data.get('is_l4_leader', 0))
    enabled = int(data.get('enabled', 1))
    password = data.get('password', '123456')
    email = (data.get('email') or '').strip() or None

    db.execute('''INSERT INTO users (username, password, display_name, role, roles, dept_l3, dept_l4, is_l3_leader, is_l4_leader, enabled, email)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
               (username, generate_password_hash(password), display_name, role, roles, dept_l3, dept_l4, is_l3_leader, is_l4_leader, enabled, email))
    db.commit()
    return jsonify({'success': True, 'id': db.execute("SELECT last_insert_rowid()").fetchone()[0]})


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@token_required
@admin_required
def admin_update_user(user_id):
    """管理员编辑用户信息"""
    data = request.get_json()
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    # Allowed fields to update
    allowed = ['role', 'roles', 'enabled', 'display_name', 'dept_l3', 'dept_l4',
               'l3_leader', 'l4_leader', 'is_l3_leader', 'is_l4_leader',
               'email']
    updates = []
    params = []
    for k in allowed:
        if k in data:
            updates.append(f"{k} = ?")
            params.append(data[k])

    if 'password' in data and data['password']:
        updates.append("password = ?")
        params.append(generate_password_hash(data['password']))

    if not updates:
        return jsonify({'error': '无更新字段'}), 400

    params.append(user_id)
    db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
    db.commit()
    return jsonify({'success': True})


@app.route('/api/admin/users/<int:user_id>/reset-password', methods=['POST'])
@token_required
@admin_required
def admin_reset_password(user_id):
    """管理员重置用户密码，返回新密码明文"""
    db = get_db()
    user = db.execute("SELECT id, username, display_name FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    new_password = secrets.token_urlsafe(6)  # 生成约8位随机密码
    db.execute("UPDATE users SET password = ? WHERE id = ?",
               (generate_password_hash(new_password), user_id))
    db.commit()
    return jsonify({
        'success': True,
        'new_password': new_password,
        'message': f'用户 {user["display_name"]} 的密码已重置为：{new_password}'
    })


@app.route('/api/admin/users/import-roster', methods=['POST'])
@token_required
@admin_required
def admin_import_roster():
    """从花名册重新导入用户"""
    from import_roster import import_roster
    added, updated, errors = import_roster()
    return jsonify({
        'added': added,
        'updated': updated,
        'errors': errors,
    })


@app.route('/api/admin/repair-checkin-stages', methods=['POST'])
@token_required
@admin_required
def admin_repair_checkin_stages():
    """一次性修复：将主表 checkin_stage 同步为最新 followup_log 的阶段值"""
    db = get_db()
    # 找出所有有填报记录的客户，用最新一次填报的 checkin_stage 覆盖主表
    repaired = db.execute("""
        UPDATE key_account_hardware h
        SET checkin_stage = (
            SELECT fl.checkin_stage FROM key_account_followup_logs fl
            WHERE fl.hardware_id = h.id AND fl.checkin_stage IS NOT NULL AND fl.checkin_stage != ''
            ORDER BY fl.id DESC LIMIT 1
        ),
        updated_at = datetime('now','localtime')
        WHERE EXISTS (
            SELECT 1 FROM key_account_followup_logs fl
            WHERE fl.hardware_id = h.id AND fl.checkin_stage IS NOT NULL AND fl.checkin_stage != ''
        )
        AND (h.checkin_stage IS NULL OR h.checkin_stage = ''
             OR h.checkin_stage != (
                 SELECT fl2.checkin_stage FROM key_account_followup_logs fl2
                 WHERE fl2.hardware_id = h.id AND fl2.checkin_stage IS NOT NULL AND fl2.checkin_stage != ''
                 ORDER BY fl2.id DESC LIMIT 1
             ))
    """).rowcount
    db.commit()
    return jsonify({'message': f'已修复 {repaired} 条记录的 checkin_stage', 'repaired': repaired})


@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory(app.static_folder, path)


# ============================================================
# 模块级初始化（gunicorn + local 都会执行）
# ============================================================
init_db()
init_key_account_db()
init_stages_config()
init_followup_logs()
# 迁移：添加新字段（如果列不存在则添加）
_db = sqlite3.connect(DB_PATH)
for _col, _type in [
    ('opportunity_amount', 'REAL DEFAULT 0'),
    ('opportunity_status', 'TEXT'),
    ('stage4_reported', 'INTEGER DEFAULT 0'),
    ('stage4_decision_role', 'TEXT'),
    ('stage4_meeting_held', 'INTEGER DEFAULT 0'),
    ('stage4_meeting_files', 'TEXT'),
    ('stage6_no_intent_reason', 'TEXT'),
    ('created_date', 'TEXT'),
    ('lent_date', 'TEXT'),           # 借用日期（最早）
    ('purchase_date', 'TEXT'),       # 采购日期
    ('checkin_date', 'TEXT'),        # 打点日期 = MIN(借用日期, 采购日期)
]:
    try:
        _db.execute(f"ALTER TABLE key_account_hardware ADD COLUMN {_col} {_type}")
    except sqlite3.OperationalError:
        pass
_db.close()


# ============================================================
# 启动
# ============================================================

if __name__ == '__main__':
    init_db()
    init_key_account_db()
    init_stages_config()
    init_followup_logs()
    host = os.environ.get('SERVER_HOST', '0.0.0.0')
    port = int(os.environ.get('SERVER_PORT', '5100'))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    print(f"销售工作台启动: http://localhost:{port}")
    app.run(host=host, port=port, debug=debug)
